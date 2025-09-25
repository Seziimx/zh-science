from __future__ import annotations
import os
from typing import Optional, List
import sys

# Make backend package importable when running as: python -m backend.scripts.import_kokson_excel "Koksost.xlsm"
SCRIPT_DIR = os.path.dirname(__file__)
BACKEND_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, '..'))
if BACKEND_DIR not in sys.path:
    sys.path.append(BACKEND_DIR)

from app.db import SessionLocal
from app.models import Source, Publication, Author, User
from sqlalchemy import select, and_
import hashlib
from app.config import get_settings
import random
import string
import hashlib

def _normalize_doc_type(s: str | None) -> str | None:
    if not s:
        return None
    x = str(s).strip().lower()
    # Common variants → canonical labels used in frontend filters (Kazakh)
    book_variants = {
        'book','books','книга','книги','кітап','кітаптар','монография (книга)','учебник','учебное пособие','оқу-әдістемелік құрал'
    }
    conf_variants = {
        'conference','conf','конференция','конференции','сборник конференции','конференциялар жинағы','proceedings'
    }
    if x in book_variants:
        return 'Кітаптар'
    if x in conf_variants:
        return 'Конференциялар жинағы'
    # Book subtypes normalization (all map to a single canonical label)
    book_subtypes = {
        'оқу-әдістемелік құрал': 'Оқу-әдістемелік құрал',
        'оқу құралы': 'Оқу-әдістемелік құрал',
        'оқу қуралы': 'Оқу-әдістемелік құрал',
        'оқулық': 'Оқу-әдістемелік құрал',
    }
    if x in book_subtypes:
        return book_subtypes[x]
    # Other book types provided by user
    other_book = {
        'танымдық жинақ': 'Танымдық жинақ',
        'энциклопедия': 'Энциклопедия',
    }
    if x in other_book:
        return other_book[x]
    # Conference subtypes from user data → normalize to consistent Kazakh labels used in filters
    conf_subtypes = {
        # International
        'халықаралық': 'Халықаралық',
        'международный': 'Халықаралық',
        # Foreign
        'шетелдік': 'Шетелдік',
        'иностранных': 'Шетелдік',
        'иностранец': 'Шетелдік',
        # Republican
        'республикалық': 'Республикалық',
        'республиканец': 'Республикалық',
        # Regional
        'аймақтық': 'Аймақтық',
        'аимактык': 'Аймақтық',
        'аймактык': 'Аймақтық',
        # Intra-university
        'университетішілік': 'Университетішілік',
        'университетский': 'Университетішілік',
    }
    if x in conf_subtypes:
        return conf_subtypes[x]
    # Leave as-is (but keep original case provided later)
    return s
def _gen_password_hash(pw: str) -> str:
    try:
        # prefer werkzeug if installed
        from werkzeug.security import generate_password_hash as _wgh  # type: ignore
        return _wgh(pw)
    except Exception:
        # fallback to sha256 with salt
        salt = get_settings().PASSWORD_SALT
        return hashlib.sha256((salt + pw).encode("utf-8")).hexdigest()

def _rand_password(length: int = 5) -> str:
    alphabet = string.ascii_letters + string.digits
    return ''.join(random.choice(alphabet) for _ in range(length))

def _deterministic_password_from_name(full_name: str, length: int = 6) -> str:
    # Normalize name and build SHA-256, then map to base36 uppercase
    name = (full_name or '').replace('\u00A0',' ').strip().upper()
    if not name:
        name = 'USER'
    digest = hashlib.sha256(name.encode('utf-8')).digest()
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    # Map first N bytes to alphabet indices to ensure stable password
    chars = [alphabet[digest[i] % len(alphabet)] for i in range(length)]
    return ''.join(chars)

def _clean(s: Optional[object]) -> str:
    if s is None:
        return ''
    txt = str(s).replace('\xa0', ' ').strip()
    return '' if txt.lower() == 'nan' else txt

def import_kokson_from_excel(db, path: str) -> dict:
    from openpyxl import load_workbook
    if not os.path.exists(path):
        print(f"Kokson Excel not found: {path}")
        return {"created": 0, "updated": 0, "skipped": 0}

    wb = load_workbook(path, data_only=True)
    # Prefer a publications-like sheet if present, else active
    sheet_names = [s.lower() for s in wb.sheetnames]
    ws = wb.active
    for name in wb.sheetnames:
        ln = name.lower()
        if any(k in ln for k in ['publication', 'publications', 'стать', 'журнал', 'book', 'конферен', 'conf']):
            ws = wb[name]
            break
    # Detect header row within first 15 rows
    def _row_vals(r):
        return [str((c.value or '')).strip() for c in r[0:ws.max_column]]
    header_row_ix = 1
    header = _row_vals(next(ws.iter_rows(min_row=1, max_row=1)))
    # scan rows 1..15 to find a row that contains any of the expected aliases for title and year
    expected_title_aliases = {'name','название','наименование','название статьи','название публикации','название книги','заглавие','наименование публикации','название работы','наименование труда','атауы','атауы (еңбек)','труд','наименование документа'}
    expected_year_aliases = {'year','год','год издания','year of publication','год публикации','год выхода','жылы','шығарылған жыл'}
    for ridx, r in enumerate(ws.iter_rows(min_row=1, max_row=15), start=1):
        vals = [str((c.value or '')).strip().lower() for c in r[0:ws.max_column]]
        has_title = any(v in expected_title_aliases for v in vals)
        has_year = any(v in expected_year_aliases for v in vals)
        if has_title and has_year:
            header_row_ix = ridx
            header = _row_vals(r)
            break
    idx = {h.lower(): i for i, h in enumerate(header)}
    def col(*alts: str) -> int | None:
        for a in alts:
            if a.lower() in idx: return idx[a.lower()]
        return None

    # Broadened header variants to accommodate Science Book/Conference formats
    c_type = col('Type','Тип','Тип публикации','Вид публикации')
    c_name = col(
        'Name','Название','Наименование',
        'Название статьи','Название публикации','Название книги','Заглавие','Наименование публикации',
        'Название работы','Наименование труда','Атауы','Атауы (еңбек)','Труд','Наименование документа',
        'namebook'
    )
    c_rname = col(
        'Rname','Источник','Журнал','Издание','Издательство','Сборник','Конференция','Источник/Издание',
        'Сборник материалов','Материалы конференции','Баспа','Басылым','Жарияланым', 'Издатель'
    )
    c_year = col('Year','Год','Год издания','Year of publication','Год публикации','Год выхода','Жылы','Шығарылған жыл')
    c_sany = col('Sany')
    c_page = col('Page','Страницы','Стр.','Страница','Pages','pagecount')
    c_author = col(
        'Author','author','Authors','Author(s)','Авторы','Авторы (ФИО)','Автор(ы)','Авторы статьи','Авторлар',
        'Авторы (ФИО полностью)','Авторы (полностью)','Докладчик','Докладчики','Спикер','Спикеры','Баяндамашы','Докладшы',
        'Тезис авторы','Автор тезиса','Presenter','Presenters'
    )
    c_co = col(
        'Соавтор','Соавторы','Соавтор(ы)','Соавторы тезиса','Соавт.','Соавторлар','Бірлескен авторлар','Қоса авторлар',
        'Coauthor','Co-authors','coauthor','co-authors'
    )
    c_pass = col('Пароль','Password')
    c_link = col('Ссылка','URL','Ссылка на источник','Ссылка (URL)')
    c_pdf = col('PDF','Файл','ПДФ','Скан','file')
    c_date = col('data','Дата','Дата публикации','Дата выхода')
    c_status = col('Статус','Status','Статус публикации')
    c_note = col('Комментарий','Ескерту','Примечание','Comment','Note','stонеов')
    c_lang = col('Язык','Language','Язык публикации')
    c_login = col('Login','Логин')
    c_issn = col('ISSN','eISSN','ISSN/eISSN')
    # Additional single-purpose columns for fallback source construction
    c_city = col('city','город','қала')
    c_baspa = col('baspa','баспа','издательство','publisher')
    c_place = col('otkenjer','өткен жер','otken jeri','место проведения')

    created = 0
    updated = 0
    skipped = 0
    skipped_reasons: dict[str, int] = {}
    skipped_examples: dict[str, list[str]] = {}
    def _skip(reason: str, example: str | None = None):
        nonlocal skipped
        skipped += 1
        skipped_reasons[reason] = skipped_reasons.get(reason, 0) + 1
        if example:
            arr = skipped_examples.setdefault(reason, [])
            if len(arr) < 10:
                arr.append(example)
    src_created = 0
    src_reused = 0
    base_pdf = "http://science.arsu.kz/"

    # Announce detected context
    kind = 'KOKSON'
    fname = os.path.basename(path).lower()
    if 'authorship' in fname:
        kind = 'AUTHORSHIP'
    elif 'book' in fname:
        kind = 'BOOK'
    elif 'konfer' in fname or 'conf' in fname or 'conference' in fname:
        kind = 'KONFERENSIA'
    print(f"[import] Publications sheet detected: {kind}")

    data_start = header_row_ix + 1
    for r in ws.iter_rows(min_row=data_start, values_only=True):
        def val(c):
            return _clean(r[c]) if (c is not None and c < len(r) and r[c] is not None) else ''
        title = val(c_name)
        if not title:
            _skip('no_title'); continue
        doc_type_raw = val(c_type) or None
        doc_type = _normalize_doc_type(doc_type_raw)
        if not doc_type:
            # Infer from filename when Type column is absent
            fname = os.path.basename(path).lower()
            if 'book' in fname:
                doc_type = _normalize_doc_type('book')
            elif 'konfer' in fname or 'conf' in fname or 'conference' in fname:
                doc_type = _normalize_doc_type('conference')
            elif 'author' in fname or 'article' in fname:
                doc_type = _normalize_doc_type('article')
        src_name = val(c_rname) or None
        if not src_name:
            city = val(c_city) if c_city is not None else ''
            baspa = val(c_baspa) if c_baspa is not None else ''
            place = val(c_place) if c_place is not None else ''
            if baspa or city:
                src_name = ": ".join([s for s in [city, baspa] if s])
            elif place:
                src_name = f"Конференция: {place}"
        issn = val(c_issn) or None
        year_txt = val(c_year)
        date_txt = val(c_date)
        raw_lang = val(c_lang)
        lng = (raw_lang or '').strip().lower()
        if lng in ("ru","rus","russian","рус","русский"):
            language = 'ru'
        elif lng in ("kz","kaz","kazakh","қаз","қазақша","каз","казахский","қазақ тілі"):
            language = 'kz'
        elif lng in ("en","eng","english","анг","английский"):
            language = 'en'
        else:
            # treat 'общ'/'общий'/'—' as unknown
            if lng in ("общ","общий","-","—","н/д","none","unknown","unk"):
                language = None
            else:
                language = (raw_lang or None)
        st_norm = (val(c_status) or '').strip().lower()
        # Normalize broader set of statuses
        approved_set = {'одобрено','одобрен','approved','қабылданды','accepted'}
        rejected_set = {'отклонено','отклонен','rejected','қабылданбады','қайтарылды','return','declined'}
        status = 'approved'
        if st_norm:
            if st_norm in approved_set:
                status = 'approved'
            elif st_norm in rejected_set:
                status = 'rejected'
        # capture optional note/comment
        raw_note = val(c_note) if c_note is not None else ''
        url = val(c_link) or None
        pdf_val = val(c_pdf) or None
        authors_main = val(c_author)
        authors_co = val(c_co)
        # If both empty -> skip. If main empty but coauthors exist -> keep as coauthors only (main_authors_count=0 later)
        if not authors_main and not authors_co:
            _skip('no_main_author', title); continue
        login = val(c_login)
        pwd = val(c_pass)
        sany = val(c_sany)
        page = val(c_page)

        # Parse date/year (support datetime cell or dd.mm.yyyy string)
        pub_date = None
        _year = None
        try:
            from datetime import date, datetime
            raw = r[c_date] if (c_date is not None and c_date < len(r)) else None
            if isinstance(raw, (date, datetime)):
                pub_date = date(raw.year, raw.month, raw.day)
                _year = raw.year
        except Exception:
            pass
        if date_txt and not _year:
            try:
                d,m,y = date_txt.split('.')
                from datetime import date as _d
                pub_date = _d(int(y), int(m), int(d))
                _year = int(y)
            except Exception:
                pass
        if not _year:
            try:
                # handle float-like '2021.0' and strings with trailing decimals
                yraw = year_txt
                if isinstance(yraw, str) and yraw.endswith('.0'):
                    yraw = yraw.split('.')[0]
                _year = int(float(yraw))
            except Exception:
                _year = None
        if not _year and year_txt:
            try:
                import re as _re
                m = _re.search(r"\b(19\d{2}|20\d{2}|21\d{2})\b", str(year_txt))
                if m:
                    _year = int(m.group(1))
            except Exception:
                pass
        if not _year:
            _skip('no_year', f"{title} | {year_txt} | {date_txt}"); continue

        # Source upsert
        src = None
        if issn:
            src = db.execute(select(Source).where(Source.issn==issn)).scalars().first()
        if not src and src_name:
            src = db.execute(select(Source).where(Source.name==src_name)).scalars().first()
        if src:
            src_reused += 1
        if not src:
            # Heuristic for source type
            src_type = 'journal'
            if doc_type and str(doc_type).lower().startswith('conf'):
                src_type = 'conference'
            elif src_name and any(k in src_name.lower() for k in ['конферен', 'conference', 'symposium']):
                src_type = 'conference'
            src = Source(name=src_name or (issn or 'Unknown'), issn=issn, type=src_type)
            db.add(src); db.flush(); src_created += 1

        # Authors build
        # Support ';', newlines and commas as separators
        import re as _re
        def _split_names(s: str) -> list[str]:
            parts = _re.split(r"[;\n,]+", s or '')
            return [p.strip() for p in parts if p and p.strip()]
        main_list = _split_names(authors_main)
        co_list = _split_names(authors_co)
        main_missing_but_co = (len(main_list) == 0 and len(co_list) > 0)
        combined = main_list + co_list
        author_objs: list[Author] = []
        for nm in combined:
            ex = db.execute(select(Author).where(Author.display_name==nm)).scalars().first()
            if ex:
                author_objs.append(ex)
            else:
                a = Author(display_name=nm, normalized_name=' '.join(nm.lower().split()))
                db.add(a); db.flush(); author_objs.append(a)

        # Optional link to user: prefer login; fallback to exact full_name match
        user_id = None
        # If login missing, use first author as login/password suggestion
        if not login and main_list:
            login = main_list[0]
        # Deterministic initial password from main author's name (4 chars A-Z0-9)
        main_full_name = (main_list[0] if main_list else login) or ''
        det_pw = _deterministic_password_from_name(main_full_name, 6)
        u = None
        if login:
            u = db.execute(select(User).where(User.login==login)).scalar_one_or_none()
        if not u and main_list:
            # Try exact case-insensitive full_name match
            from sqlalchemy import func
            full_name = main_list[0]
            u = db.execute(select(User).where(func.lower(User.full_name)==func.lower(full_name))).scalar_one_or_none()
        if not u and main_list and login:
            # Create a new user if nothing matched
            full_name = main_list[0]
            u = User(full_name=full_name, login=login, email=None, role='teacher', faculty='', department='', position='', degree='', password_hash=_gen_password_hash(det_pw), initial_password=det_pw, created_source='import')
            db.add(u); db.flush()
        elif u:
            # Set deterministic password only if user has no initial_password yet
            if not getattr(u, 'initial_password', None):
                u.initial_password = det_pw
                try:
                    u.password_hash = _gen_password_hash(det_pw)
                except Exception:
                    pass
        if u:
            user_id = u.id

        # Build pdf URL
        pdf_url = None
        if pdf_val:
            v = str(pdf_val).lstrip()
            if v.startswith('http'):
                pdf_url = v
            elif v.startswith('/files/'):
                pdf_url = 'http://science.arsu.kz' + v
            elif v.startswith('files/'):
                pdf_url = 'http://science.arsu.kz/' + v
            else:
                pdf_url = base_pdf + v

        # Upsert publication by (title, year, source)
        existing = db.execute(select(Publication).where(and_(Publication.title==title, Publication.year==_year, Publication.source_id==(src.id if src else None)))).scalars().first()
        note_parts = []
        # Intentionally ignore Sany/Page/Пароль in notes per requirements
        if raw_note: note_parts.append(raw_note)
        # If there is a comment but status is unknown, treat as rejected with note
        if raw_note and st_norm == '':
            status = 'rejected'
        # Excel logic: if URL is empty and PDF present -> "Надо вставить ссылку на источник"
        # if PDF empty and URL present -> "Надо добавить ПДФ"
        # We also mark such rows as rejected to expose them в админке.
        missing_msgs: list[str] = []
        if (url is None or url == '') and (pdf_val is not None and pdf_val != ''):
            missing_msgs.append("Надо вставить ссылку на источник")
        if (pdf_val is None or pdf_val == '') and (url is not None and url != ''):
            missing_msgs.append("Надо добавить ПДФ")
        if missing_msgs:
            note_parts.extend(missing_msgs)
            status = 'rejected'
        note_val = '; '.join(note_parts) if note_parts else None

        if existing:
            p = existing
            p.doc_type = doc_type or p.doc_type
            p.language = language or p.language
            p.url = url or p.url
            p.pdf_url = pdf_url or p.pdf_url
            p.published_date = pub_date or p.published_date
            p.status = status or p.status
            p.note = note_val or p.note
            p.main_authors_count = (0 if main_missing_but_co else (len(main_list) or None))
            p.authors = author_objs
            if user_id: p.user_id = user_id
            updated += 1
        else:
            p = Publication(
                title=title,
                year=_year,
                doi=None,
                pdf_url=pdf_url,
                url=(url or None),
                citations_count=0,
                source=src,
                status=status,
                uploader_id='import',
                uploaded_by_role='admin',
                language=language,
                upload_source='kokson',
                doc_type=doc_type,
                published_date=pub_date,
                main_authors_count=(0 if main_missing_but_co else (len(main_list) or None)),
                user_id=user_id,
            )
            db.add(p); db.flush(); p.authors = author_objs
            created += 1

    if created or updated:
        db.commit()
    print(f"[import] Sources imported/updated: created={src_created}, reused={src_reused}")
    print(f"[import] Publications imported: created={created}, updated={updated}, skipped={skipped}")
    if skipped:
        print("[import] Skipped breakdown:")
        for k, v in skipped_reasons.items():
            print(f"  - {k}: {v}")
            ex = skipped_examples.get(k) or []
            for s in ex:
                print(f"      example: {s}")
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "src_created": src_created,
        "src_reused": src_reused,
        "skipped_reasons": skipped_reasons,
        "skipped_examples": {k: (v[:5] if isinstance(v, list) else v) for k, v in skipped_examples.items()},
    }

def main():
    # Allow passing multiple Excel paths via CLI; if none provided, try common candidates in repo root.
    args = sys.argv[1:]
    paths: list[str] = []
    if args:
        for a in args:
            p = os.path.abspath(a)
            if os.path.exists(p):
                paths.append(p)
            else:
                print(f"File not found: {a}")
    if not paths:
        candidates = [
            os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', 'Koksost.xlsm')),
            os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', 'Science Authorship (All).xlsx')),
            os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', 'Science Book (All).xlsx')),
            os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', 'Science Konferensia (All).xlsx')),
        ]
        for p in candidates:
            if os.path.exists(p):
                paths.append(p)
    if not paths:
        print("No Excel files found (provide paths or place files in repo root)")
        return
    db = SessionLocal()
    try:
        totals = {"created": 0, "updated": 0, "skipped": 0}
        for path in paths:
            base = os.path.basename(path)
            # Label per file
            label = 'Kokson'
            bl = base.lower()
            if 'authorship' in bl:
                label = 'Science Authorship'
            elif 'book' in bl:
                label = 'Science Book'
            elif 'konfer' in bl or 'conf' in bl or 'conference' in bl:
                label = 'Science Konferensia'
            print(f"[import] {label} file: {base}")
            res = import_kokson_from_excel(db, path)
            try:
                totals["created"] += int(res.get("created", 0))
                totals["updated"] += int(res.get("updated", 0))
                totals["skipped"] += int(res.get("skipped", 0))
            except Exception:
                pass
            # Per-file summary echo similar to example
            print(f"{label} import: created={res.get('created',0)}, updated={res.get('updated',0)}, skipped={res.get('skipped',0)}")
        print(f"[import] All files done: files={len(paths)} created={totals['created']} updated={totals['updated']} skipped={totals['skipped']}")
    finally:
        db.close()

if __name__ == '__main__':
    main()
