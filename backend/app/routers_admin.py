 
from __future__ import annotations
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, Body, UploadFile, File, Header, Response
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session, joinedload, aliased
from sqlalchemy import select, desc, asc, or_, func
import os
import json
import hashlib

from .db import get_db
from .config import get_settings
from .models import Publication, User, Author, publication_authors
from .schemas import PublicationOut, UserOut, UserWithCountOut, UserCreate, UserUpdate, UserPasswordChange, MatchPreviewResponse, LoginCheckResponse

router = APIRouter(prefix="/admin", tags=["admin"])

UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'uploads'))


def require_admin(authorization: Optional[str] = Header(default=None)) -> None:
    settings = get_settings()
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    parts = authorization.split()
    token = parts[-1] if parts else authorization
    if token != settings.ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Invalid admin token")

# -----------------------------
# Helpers (password hash, name variants)
# -----------------------------
def _hash_password(pw: str, salt: str) -> str:
    data = (salt + pw).encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def _deterministic_pw(full_name: str, length: int = 6) -> str:
    """Generate deterministic A–Z0–9 password of fixed length from full name."""
    name = (full_name or '').replace('\u00A0',' ').strip().upper() or 'USER'
    digest = hashlib.sha256(name.encode('utf-8')).digest()
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    return ''.join(alphabet[digest[i] % len(alphabet)] for i in range(length))


def _generate_name_variants(full_name: str) -> List[str]:
    parts = [p for p in full_name.replace("\u00A0", " ").strip().split() if p]
    if not parts:
        return []
    # assume first token is last name in local formatting (common in CIS)
    last = parts[0]
    initials = "".join([(p[0] + ".") for p in parts[1:]])  # S.S.
    one_init = (parts[1][0] + ".") if len(parts) > 1 else ""

    def translit(s: str) -> str:
        table = {
            "А":"A","Б":"B","В":"V","Г":"G","Д":"D","Е":"E","Ё":"E","Ж":"Zh","З":"Z","И":"I","Й":"Y","К":"K","Л":"L","М":"M","Н":"N","О":"O","П":"P","Р":"R","С":"S","Т":"T","У":"U","Ф":"F","Х":"Kh","Ц":"Ts","Ч":"Ch","Ш":"Sh","Щ":"Sch","Ы":"Y","Э":"E","Ю":"Yu","Я":"Ya",
            "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"zh","з":"z","и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o","п":"p","р":"r","с":"s","т":"t","у":"u","ф":"f","х":"kh","ц":"ts","ч":"ch","ш":"sh","щ":"sch","ы":"y","э":"e","ю":"yu","я":"ya",
        }
        return "".join([table.get(ch, ch) for ch in s])

    candidates = set()
    variants_last = {last, translit(last)}
    for l in variants_last:
        candidates.add(l)
        if one_init:
            candidates.add(f"{l} {one_init}")
        if initials:
            candidates.add(f"{l} {initials}")
        if one_init:
            candidates.add(f"{one_init} {l}")
        if initials:
            spaced = " ".join([f"{c}." for c in initials.replace('.', '') if c]) if initials else ""
            if spaced:
                candidates.add(f"{spaced} {l}")
    if initials:
        candidates.add(initials)
    return sorted({c.strip() for c in candidates if c.strip()})


# -----------------------------
# Publications moderation
# -----------------------------
@router.get("/publications", response_model=list[PublicationOut])
def list_publications(
    status: Optional[str] = Query(default=None, description="pending|approved|rejected"),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=1, le=200),
    order: str = Query(default="created_desc", description="created_desc|year_desc|year_asc"),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    stmt = select(Publication).options(joinedload(Publication.source), joinedload(Publication.authors))
    if status:
        stmt = stmt.where(Publication.status == status)
    if order == "year_asc":
        stmt = stmt.order_by(asc(Publication.year), Publication.id)
    elif order == "year_desc":
        stmt = stmt.order_by(desc(Publication.year), Publication.id)
    else:
        stmt = stmt.order_by(desc(Publication.created_at))
    offset = (page - 1) * per_page
    stmt = stmt.offset(offset).limit(per_page)
    rows = db.execute(stmt).scalars().unique().all()
    return [PublicationOut.model_validate(p) for p in rows]


@router.post("/publications/{pub_id}/approve")
def approve_publication(pub_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    pub = db.get(Publication, pub_id)
    if not pub:
        raise HTTPException(status_code=404, detail="Publication not found")
    pub.status = "approved"
    pub.note = None
    db.add(pub)
    db.commit()
    return {"ok": True}

# -----------------------------
# Users export (login + initial password)
# -----------------------------
@router.get("/users/export")
def export_users(
    fmt: str = Query(default="xlsx", description="xlsx|csv"),
    created_source: Optional[str] = Query(default=None, description="admin|import|api"),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    from fastapi.responses import Response
    from sqlalchemy import select
    users = db.execute(select(User)).scalars().all()
    if created_source:
        users = [u for u in users if (u.created_source or '') == created_source]
    headers_row = [
        "ID","ФИО","Логин","Начальный пароль","Роль","Факультет","Кафедра","Должность","Степень","Email","Активен"
    ]
    rows = []
    for u in users:
        rows.append([
            u.id,
            u.full_name or '',
            u.login or '',
            getattr(u, 'initial_password', None) or '',
            u.role or '',
            u.faculty or '',
            u.department or '',
            u.position or '',
            u.degree or '',
            u.email or '',
            1 if (u.active or 0) else 0,
        ])
    if fmt == "csv":
        import io, csv
        buf = io.StringIO(); w = csv.writer(buf)
        w.writerow(headers_row); w.writerows(rows)
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=users.csv"},
        )
    try:
        from openpyxl import Workbook
        import io
        wb = Workbook(); ws = wb.active
        ws.append(headers_row)
        for r in rows: ws.append(r)
        bio = io.BytesIO(); wb.save(bio); bio.seek(0)
        return Response(
            content=bio.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=users.xlsx"},
        )
    except Exception:
        import io, csv
        buf = io.StringIO(); w = csv.writer(buf)
        w.writerow(headers_row); w.writerows(rows)
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=users.csv"},
        )


# -----------------------------
# Users management
# -----------------------------

@router.get("/users", response_model=List[UserWithCountOut])
def list_users(
    q: Optional[str] = Query(default=None),
    role: Optional[str] = Query(default=None),
    faculty: Optional[str] = Query(default=None),
    created_source: Optional[str] = Query(default=None, description="admin|import|api"),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
    response: Response = None,
):
    """List users with publications_count computed by strict author-name matching.
    Strategy per user: strict normalized equality to Author.display_name only.
    Counts are DISTINCT publication ids.
    """

    # Auto-backfill initial_password for users missing it (no button needed)
    settings = get_settings()
    changed = 0
    users_all = db.execute(select(User)).scalars().all()
    for u in users_all:
        ipw = getattr(u, 'initial_password', None) or ''
        if not ipw:
            pw = _deterministic_pw(u.full_name)
            u.initial_password = pw
            try:
                u.password_hash = _hash_password(pw, settings.PASSWORD_SALT)
            except Exception:
                pass
            db.add(u)
            changed += 1
    if changed:
        db.commit()

    # base users query with filters
    u_stmt = select(User)
    if q:
        like = f"%{q}%"
        u_stmt = u_stmt.where(or_(User.full_name.ilike(like), User.login.ilike(like), User.email.ilike(like)))
    if role:
        u_stmt = u_stmt.where(User.role == role)
    if faculty:
        u_stmt = u_stmt.where(User.faculty == faculty)
    if created_source:
        u_stmt = u_stmt.where(User.created_source == created_source)
    u_stmt = u_stmt.order_by(User.full_name)
    users = db.execute(u_stmt).scalars().all()

    def norm_expr(col):
        # lower(replace(replace(replace(replace(col, ',', ''), NBSP, ' '), '.', ''), ' ', ''))
        expr = func.replace(col, ",", "")
        expr = func.replace(expr, "\u00A0", " ")
        expr = func.replace(expr, ".", "")
        expr = func.replace(expr, " ", "")
        return func.lower(expr)

    out: List[UserWithCountOut] = []
    total = len(users)
    start = (page - 1) * per_page
    end = start + per_page
    page_items = users[start:end]

    for u in page_items:
        pubs_count = 0
        names = [(u.full_name or "").replace("\u00A0"," ").strip()]
        conds = []
        from sqlalchemy import or_ as _or
        for nm in names:
            if not nm:
                continue
            conds.append(Author.display_name == nm)
        if conds:
            var_count_stmt = (
                select(func.count(func.distinct(Publication.id)))
                .select_from(Publication)
                .join(publication_authors, publication_authors.c.publication_id == Publication.id)
                .join(Author, Author.id == publication_authors.c.author_id)
                .where(_or(*conds))
            )
            pubs_count = db.execute(var_count_stmt).scalar() or 0
        d = UserOut.model_validate(u).model_dump()
        d["publications_count"] = int(pubs_count)
        out.append(UserWithCountOut(**d))

    if response is not None:
        response.headers["X-Total-Count"] = str(total)
    return out


@router.post("/users", response_model=UserOut)
def create_user(payload: UserCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    """Create a new user (admin only). Ensures unique login and stores a password hash.
    If login is missing, derive from full_name by removing spaces and lowercasing (fallback adds digits if occupied).
    """
    settings = get_settings()
    full_name = (payload.full_name or '').replace('\u00A0',' ').strip()
    if not full_name:
        raise HTTPException(status_code=400, detail="full_name is required")
    # Prepare login
    desired_login = (payload.login or full_name).replace('\u00A0',' ').strip().lower().replace(' ', '').replace(',', '').replace('.', '')
    if not desired_login:
        desired_login = "user"
    # Ensure uniqueness
    i = 0
    final_login = desired_login
    from sqlalchemy import func
    def exists_login(val: str) -> bool:
        return db.execute(select(User).where(func.lower(User.login) == func.lower(val))).scalar_one_or_none() is not None
    while exists_login(final_login):
        i += 1
        final_login = f"{desired_login}{i}"
        if i > 9999:
            raise HTTPException(status_code=409, detail="Cannot allocate unique login")

    # Hash password and set initial_password (store deterministic if not provided?)
    password_plain = (payload.password or '').strip()
    if not password_plain:
        raise HTTPException(status_code=400, detail="password is required")
    pw_hash = _hash_password(password_plain, settings.PASSWORD_SALT)

    # Create user
    u = User(
        full_name=full_name,
        login=final_login,
        email=(payload.email or None),
        role=(payload.role or 'teacher'),
        faculty=payload.faculty,
        department=payload.department,
        position=payload.position,
        degree=payload.degree,
        active=1,
        password_hash=pw_hash,
        initial_password=password_plain,
        created_source='admin',
    )
    # Optionally populate name_variants
    try:
        variants = _generate_name_variants(full_name)
        if variants:
            import json
            u.name_variants = json.dumps(variants, ensure_ascii=False)
    except Exception:
        pass
    db.add(u); db.commit(); db.refresh(u)
    return UserOut.model_validate(u)


@router.post("/users/backfill_passwords")
def backfill_passwords(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Backfill deterministic initial_password for users who miss it.
    Returns number of users updated. Idempotent.
    """
    settings = get_settings()
    users = db.execute(select(User)).scalars().all()
    changed = 0
    for u in users:
        ipw = getattr(u, 'initial_password', None) or ''
        if not ipw:
            pw = _deterministic_pw(u.full_name)
            u.initial_password = pw
            try:
                u.password_hash = _hash_password(pw, settings.PASSWORD_SALT)
            except Exception:
                pass
            db.add(u)
            changed += 1
    if changed:
        db.commit()
    return {"updated": changed}


@router.post("/maintenance/normalize_languages")
def normalize_languages(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Normalize Publication.language values to ru/kz/en where possible.
    Maps common variants like 'Казахский', 'қазақша', 'kazakh' -> 'kz';
    'Русский','russian','rus' -> 'ru'; 'English','английский','eng' -> 'en'.
    Returns number of rows updated.
    """
    from sqlalchemy import update
    updated = 0
    # Build CASE-like updates per group
    variants_kz = [
        'kz','kaz','kazakh','қаз','қазақша','каз','казахский','қазақ тілі','kazakh language'
    ]
    variants_ru = [
        'ru','rus','russian','рус','русский','russian language'
    ]
    variants_en = [
        'en','eng','english','анг','английский','english language'
    ]
    # Helper: lower(trim(language)) IN variants -> set to code
    def apply(code: str, variants: list[str]):
        nonlocal updated
        stmt = (
            update(Publication)
            .where(Publication.language.is_not(None))
            .where(func.lower(func.trim(Publication.language)).in_(variants))
            .values(language=code)
        )
        res = db.execute(stmt)
        updated += res.rowcount or 0
    apply('kz', variants_kz)
    apply('ru', variants_ru)
    apply('en', variants_en)
    db.commit()
    return {"normalized": int(updated)}

@router.get("/users/match_preview", response_model=MatchPreviewResponse)
def users_match_preview(full_name: str, exact: bool = Query(default=False), db: Session = Depends(get_db), _=Depends(require_admin)):
    """Preview publications matched to a person's full name.
    If exact=True or by default, use strict normalized equality only.
    """
    norm = full_name.replace("\u00A0", " ").strip()
    if not norm:
        return MatchPreviewResponse(count=0, examples=[], publications=[])

    def normalized_expr(col):
        # lower(trim(replace(non-breaking space, space)))
        return func.lower(func.replace(func.replace(col, ",", ""), "\u00A0", " "))

    pubs: list[Publication] = []
    ids: set[int] = set()

    # 1) Exact normalized equality match Author.display_name == full_name (case-insensitive)
    if exact:
        eq_stmt = (
            select(Publication)
            .join(publication_authors, publication_authors.c.publication_id == Publication.id)
            .join(Author, Author.id == publication_authors.c.author_id)
            .where(normalized_expr(Author.display_name) == norm.lower())
            .options(joinedload(Publication.source), joinedload(Publication.authors))
            .limit(50)
        )
        pubs = db.execute(eq_stmt).scalars().unique().all()
        if pubs:
            ids = {p.id for p in pubs}
            return MatchPreviewResponse(count=len(ids), examples=[norm], publications=[PublicationOut.model_validate(p) for p in pubs])
        # fallback strict: require last name and initials present
        parts = [p for p in norm.split() if p]
        last = parts[0] if parts else ''
        inits = ''.join([p[0] for p in parts[1:]])
        if last:
            cond = normalized_expr(Author.display_name).like(f"%{last}%")
            for ch in inits:
                cond = cond & normalized_expr(Author.display_name).like(f"%{ch}%")
            fb_stmt = (
                select(Publication)
                .join(publication_authors, publication_authors.c.publication_id == Publication.id)
                .join(Author, Author.id == publication_authors.c.author_id)
                .where(cond)
                .options(joinedload(Publication.source), joinedload(Publication.authors))
                .limit(50)
            )
            pubs = db.execute(fb_stmt).scalars().unique().all()
            if pubs:
                ids = {p.id for p in pubs}
                return MatchPreviewResponse(count=len(ids), examples=[norm], publications=[PublicationOut.model_validate(p) for p in pubs])
    # If nothing found, return empty (no fuzzy fallback in strict mode)
    return MatchPreviewResponse(count=0, examples=[norm], publications=[])


@router.post("/users", response_model=UserOut)
def create_user(payload: UserCreate, db: Session = Depends(get_db), _=Depends(require_admin)):
    settings = get_settings()
    # ensure unique login
    desired_login = (payload.login or payload.email or payload.full_name.replace(" ", ".").lower()).strip()
    if not desired_login:
        raise HTTPException(status_code=400, detail="Login is required")
    # check email uniqueness if provided
    if payload.email:
        exists_email = db.execute(select(User).where(User.email == payload.email)).scalar_one_or_none()
        if exists_email:
            raise HTTPException(status_code=400, detail="Email already exists")
    # disallow using reserved admin static login
    if desired_login == settings.ADMIN_LOGIN:
        raise HTTPException(status_code=400, detail="Login is reserved")
    login = desired_login
    if db.execute(select(User).where(User.login == login)).scalar_one_or_none():
        # ensure uniqueness
        login = f"{login}.{hashlib.sha1(login.encode()).hexdigest()[:6]}"
    variants = _generate_name_variants(payload.full_name)
    u = User(
        full_name=payload.full_name,
        email=payload.email,
        role=payload.role,
        faculty=payload.faculty,
        department=payload.department,
        position=payload.position,
        degree=payload.degree,
        login=login,
        password_hash=_hash_password(payload.password, settings.PASSWORD_SALT),
        name_variants=json.dumps(variants, ensure_ascii=False),
        active=1,
        created_source="admin",
    )
    db.add(u)
    db.flush()

    # auto-link publications by author variants
    if variants:
        conds = [Author.display_name.ilike(f"%{v}%") for v in variants]
        stmt = (
            select(Publication)
            .join(publication_authors, publication_authors.c.publication_id == Publication.id)
            .join(Author, Author.id == publication_authors.c.author_id)
            .where(or_(*conds))
        )
        pubs = db.execute(stmt).scalars().unique().all()
        for p in pubs:
            p.user_id = u.id
            db.add(p)

    db.commit()
    db.refresh(u)
    return UserOut.model_validate(u)


@router.get("/users/check_login", response_model=LoginCheckResponse)
def check_login(login: str, db: Session = Depends(get_db), _=Depends(require_admin)):
    settings = get_settings()
    exists = db.execute(select(User).where(User.login == login)).scalar_one_or_none()
    # also treat static ADMIN_LOGIN as reserved
    if login == settings.ADMIN_LOGIN:
        return LoginCheckResponse(available=False)
    return LoginCheckResponse(available=not bool(exists))


@router.patch("/users/{user_id}/active")
def set_user_active(user_id: int, active: int = Body(embed=True), db: Session = Depends(get_db), _=Depends(require_admin)):
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    u.active = 1 if int(active) else 0
    db.add(u)
    db.commit()
    return {"ok": True}


@router.get("/users/{user_id}/publications", response_model=List[PublicationOut])
def user_publications(
    user_id: int,
    match: str = Query(default="initials", description="exact | initials | broad"),
    # 'exact'    -> only exact normalized equality
    # 'initials' -> exact + last name with all initials present (recommended)
    # 'broad'    -> initials + name variants + legacy user_id link
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Return publications for a user by matching their full_name to authors.
    match parameter controls strictness: exact | initials | broad.
    """
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    def norm_expr(col):
        # lower(remove NBSP, spaces, dots, commas)
        expr = func.replace(col, "\u00A0", " ")
        expr = func.replace(expr, ",", "")
        expr = func.replace(expr, ".", "")
        expr = func.replace(expr, " ", "")
        return func.lower(expr)

    raw = (u.full_name or '').replace("\u00A0", " ").strip()
    norm = raw.lower().replace(".", "").replace(",", "").replace(" ", "")
    parts = [p for p in raw.split() if p]
    last = parts[0].lower() if parts else ''
    inits = ''.join([p[0].lower() for p in parts[1:]])

    # Build a distinct set of publication IDs
    pub_ids: set[int] = set()

    # A) exact normalized equality (always applied)
    ids_eq = db.execute(
        select(Publication.id)
        .join(publication_authors, publication_authors.c.publication_id == Publication.id)
        .join(Author, Author.id == publication_authors.c.author_id)
        .where(norm_expr(Author.display_name) == norm)
    ).all()
    pub_ids.update([row[0] for row in ids_eq])

    if match in ("initials", "broad"):
        # B) last name + all initials present
        if last:
            cond = norm_expr(Author.display_name).like(f"%{last}%")
            for ch in inits:
                cond = cond & norm_expr(Author.display_name).like(f"%{ch}%")
            ids_fb = db.execute(
                select(Publication.id)
                .join(publication_authors, publication_authors.c.publication_id == Publication.id)
                .join(Author, Author.id == publication_authors.c.author_id)
                .where(cond)
            ).all()
            pub_ids.update([row[0] for row in ids_fb])

    if match == "broad":
        # C) name variants ILIKE OR (broad)
        variants = _generate_name_variants(u.full_name or '')
        if variants:
            conds = [Author.display_name.ilike(f"%{v}%") for v in variants]
            ids_var = db.execute(
                select(Publication.id)
                .join(publication_authors, publication_authors.c.publication_id == Publication.id)
                .join(Author, Author.id == publication_authors.c.author_id)
                .where(or_(*conds))
            ).all()
            pub_ids.update([row[0] for row in ids_var])

        # D) legacy fallback: publications explicitly linked to user_id
        ids_link = db.execute(
            select(Publication.id).where(Publication.user_id == user_id)
        ).all()
        pub_ids.update([row[0] for row in ids_link])

    if not pub_ids:
        return []

    rows = db.execute(
        select(Publication)
        .where(Publication.id.in_(list(pub_ids)))
        .options(joinedload(Publication.source), joinedload(Publication.authors))
        .order_by(desc(Publication.year), Publication.id)
    ).scalars().unique().all()
    return [PublicationOut.model_validate(p) for p in rows]


@router.patch("/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, payload: UserUpdate, db: Session = Depends(get_db), _=Depends(require_admin)):
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    # handle login change with uniqueness check
    if payload.login is not None:
        new_login = payload.login.strip()
        if not new_login:
            raise HTTPException(status_code=400, detail="Login cannot be empty")
        if new_login != u.login:
            settings = get_settings()
            if new_login == settings.ADMIN_LOGIN:
                raise HTTPException(status_code=400, detail="Login is reserved")
            exists = db.execute(select(User).where(User.login == new_login)).scalar_one_or_none()
            if exists:
                raise HTTPException(status_code=400, detail="Login already exists")
            u.login = new_login
    for field in ["full_name","email","role","faculty","department","position","degree","active"]:
        val = getattr(payload, field)
        if val is not None:
            setattr(u, field, val)
    # if full_name changed, recompute variants and relink
    if payload.full_name is not None:
        variants = _generate_name_variants(u.full_name)
        u.name_variants = json.dumps(variants, ensure_ascii=False)
        # relink (simple strategy: link publications matching variants)
        conds = [Author.display_name.ilike(f"%{v}%") for v in variants]
        stmt = (
            select(Publication)
            .join(publication_authors, publication_authors.c.publication_id == Publication.id)
            .join(Author, Author.id == publication_authors.c.author_id)
            .where(or_(*conds))
        )
        pubs = db.execute(stmt).scalars().unique().all()
        for p in pubs:
            p.user_id = u.id
            db.add(p)
    db.add(u)
    db.commit()
    db.refresh(u)
    return UserOut.model_validate(u)


@router.patch("/users/{user_id}/password")
def update_user_password(user_id: int, payload: UserPasswordChange, db: Session = Depends(get_db), _=Depends(require_admin)):
    settings = get_settings()
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    u.password_hash = _hash_password(payload.password, settings.PASSWORD_SALT)
    db.add(u)
    db.commit()
    return {"ok": True}


@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(u)
    db.commit()
    return {"ok": True}
@router.post("/publications/{pub_id}/reject")
def reject_publication(
    pub_id: int,
    note: Optional[str] = Body(default=None, embed=True),  # expect {"note": "..."}
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    pub = db.get(Publication, pub_id)
    if not pub:
        raise HTTPException(status_code=404, detail="Publication not found")
    pub.status = "rejected"
    pub.note = note
    db.add(pub)
    db.commit()
    return {"ok": True}


@router.delete("/publications/{pub_id}")
def delete_publication(pub_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    pub = db.get(Publication, pub_id)
    if not pub:
        raise HTTPException(status_code=404, detail="Publication not found")
    # try remove file from disk if present
    try:
        if pub.pdf_url:
            fname = os.path.basename(pub.pdf_url)
            fpath = os.path.join(UPLOAD_DIR, fname)
            if os.path.isfile(fpath):
                os.remove(fpath)
    except Exception:
        pass
    db.delete(pub)
    db.commit()
    return {"ok": True}

# -----------------------------
# Import faculty/users from Excel (upload in browser)
# -----------------------------

@router.post("/import/faculty")
def import_faculty_excel(
    authorization: Optional[str] = Header(default=None),
    file: UploadFile = File(..., description="Excel file with faculty/users (e.g., факультет.xlsx)"),
):
    """Upload an Excel file and import faculty/department/users to DB.
    Auth: pass admin token in Authorization header (raw token or 'Bearer <token>').
    """
    require_admin(authorization)  # raises 401 if invalid

    # Save to a temporary file under uploads
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(UPLOAD_DIR, f"_faculty_import_{file.filename}")
    with open(temp_path, "wb") as out:
        out.write(file.file.read())

    # Run importer
    try:
        try:
            from scripts.import_faculty_excel import import_faculty_from_excel  # type: ignore
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Importer not available: {e}")

        res = import_faculty_from_excel(temp_path)
        return {"status": "ok", **res}
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


@router.post("/import/kokson")
def import_kokson_excel(
    authorization: Optional[str] = Header(default=None),
    file: UploadFile = File(..., description="Kokson Excel (Коксон.xlsx) with authors/title/url/language/year/journal/issn"),
):
    """Upload Kokson.xlsx and import Kokson publications (type='koks').
    Auth via `Authorization: Bearer <ADMIN_TOKEN>`.
    """
    require_admin(authorization)

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(UPLOAD_DIR, f"_kokson_import_{file.filename}")
    with open(temp_path, "wb") as out:
        out.write(file.file.read())

    try:
        try:
            from scripts.import_kokson_excel import import_kokson_from_excel  # type: ignore
            from app.db import SessionLocal
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Importer not available: {e}")

        db = SessionLocal()
        try:
            created = import_kokson_from_excel(db, temp_path)
        finally:
            db.close()
        return {"status": "ok", "kokson_imported": created}
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


@router.post("/import/publications")
def import_publications_excel(
    authorization: Optional[str] = Header(default=None),
    file: UploadFile = File(..., description="Scopus Excel with Sources/Publications (e.g., zhubanov_scopus_issn.xlsx)"),
):
    """Upload Scopus Excel and import Sources and Publications on the server.
    Auth via `Authorization: Bearer <ADMIN_TOKEN>`.
    """
    require_admin(authorization)

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(UPLOAD_DIR, f"_pubs_import_{file.filename}")
    with open(temp_path, "wb") as out:
        out.write(file.file.read())

    try:
        try:
            from scripts.import_excel import load_sources_from_excel, load_publications_from_excel  # type: ignore
            from app.db import SessionLocal
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Importer not available: {e}")

        db = SessionLocal()
        try:
            load_sources_from_excel(db, temp_path)
            created = load_publications_from_excel(db, temp_path)
        finally:
            db.close()
        return {"status": "ok", "publications_imported": created}
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


@router.patch("/publications/{pub_id}", response_model=PublicationOut)
def update_publication(
    pub_id: int,
    title: Optional[str] = None,
    year: Optional[int] = None,
    doi: Optional[str] = None,
    scopus_url: Optional[str] = None,
    pdf_url: Optional[str] = None,
    citations_count: Optional[int] = None,
    quartile: Optional[str] = None,
    percentile_2024: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    pub = db.get(Publication, pub_id)
    if not pub:
        raise HTTPException(status_code=404, detail="Publication not found")
    if title is not None:
        pub.title = title
    if year is not None:
        pub.year = year
    if doi is not None:
        pub.doi = doi
    if scopus_url is not None:
        pub.scopus_url = scopus_url
    if pdf_url is not None:
        pub.pdf_url = pdf_url
    if citations_count is not None:
        pub.citations_count = citations_count
    if quartile is not None:
        pub.quartile = quartile
    if percentile_2024 is not None:
        pub.percentile_2024 = percentile_2024
    db.add(pub)
    db.commit()
    db.refresh(pub)
    return PublicationOut.model_validate(pub)


@router.post("/maintenance/backfill-user-faculties")
def backfill_user_faculties(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Fill User.faculty based on User.department using the centralized mapping
    from app.routers_publications.DEPT_TO_FAC.
    Only updates records where faculty is empty or different from the mapped value.
    Returns counts and few examples.
    """
    try:
        # Import mapping from publications router to keep a single source of truth
        from .routers_publications import DEPT_TO_FAC  # type: ignore
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Mapping not available: {e}")

    updated = 0
    examples: list[dict] = []
    users = db.execute(select(User)).scalars().all()
    for u in users:
        dept = (u.department or '').strip()
        if not dept:
            continue
        fac_new = DEPT_TO_FAC.get(dept)
        if not fac_new:
            continue
        fac_old = (u.faculty or '').strip()
        if fac_old != fac_new:
            u.faculty = fac_new
            db.add(u)
            updated += 1
            if len(examples) < 5:
                examples.append({"user_id": u.id, "department": dept, "old": fac_old, "new": fac_new})
    if updated:
        db.commit()
    return {"created": int(created_users), "updated": int(updated), "examples": examples}


@router.get("/maintenance/unmapped-departments")
def unmapped_departments(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Return distinct User.department values that are not present in DEPT_TO_FAC mapping.
    Helps to extend the mapping with new кафедры.
    """
    try:
        from .routers_publications import DEPT_TO_FAC  # type: ignore
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Mapping not available: {e}")
    rows = db.execute(select(func.distinct(User.department)).where(User.department.isnot(None))).all()
    depts = sorted({(r[0] or '').strip() for r in rows if (r[0] or '').strip()})
    unknown = [d for d in depts if d not in DEPT_TO_FAC]
    return {"unknown_departments": unknown, "count": len(unknown)}


@router.post("/maintenance/link_kokson_publications")
def link_kokson_publications(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Link Kokson publications to users by matching author names to User.full_name.
    Strategy:
    1) Strict normalized equality (remove spaces, dots, commas, NBSP, lowercase)
    2) Fallback: last name + all initials present
    The first matching user is assigned to Publication.user_id (primary owner).
    Returns number of publications linked in this run.
    """
    def norm_expr(col):
        expr = func.replace(col, "\u00A0", " ")
        expr = func.replace(expr, ",", "")
        expr = func.replace(expr, ".", "")
        expr = func.replace(expr, " ", "")
        return func.lower(expr)

    users = db.execute(select(User)).scalars().all()
    # precompute normalized names and initials
    udata = []
    for u in users:
        raw = (u.full_name or '').replace("\u00A0", " ").strip()
        if not raw:
            continue
        norm = raw.lower().replace(".", "").replace(",", "").replace(" ", "")
        parts = [p for p in raw.split() if p]
        last = parts[0].lower() if parts else ''
        inits = ''.join([p[0].lower() for p in parts[1:]])
        udata.append({"id": u.id, "norm": norm, "last": last, "inits": inits})

    pubs = db.execute(
        select(Publication)
        .where(Publication.upload_source == "kokson")
        .where((Publication.user_id.is_(None)) | (Publication.user_id == 0))
        .limit(5000)
    ).scalars().unique().all()

    linked = 0
    for p in pubs:
        # authors text values
        authors = db.execute(
            select(Author.display_name)
            .select_from(Author)
            .join(publication_authors, publication_authors.c.author_id == Author.id)
            .where(publication_authors.c.publication_id == p.id)
        ).scalars().all()
        found_uid: int | None = None
        for a in authors:
            a_raw = (a or '').replace("\u00A0", " ").strip()
            if not a_raw:
                continue
            a_norm = a_raw.lower().replace(".", "").replace(",", "").replace(" ", "")
            # 1) strict equality
            for u in udata:
                if a_norm == u["norm"]:
                    found_uid = u["id"]
                    break
            if found_uid:
                break
            # 2) last + all initials present
            parts = [p for p in a_raw.split() if p]
            last = parts[0].lower() if parts else ''
            inits = ''.join([p[0].lower() for p in parts[1:]])
            if last:
                for u in udata:
                    if last in u["norm"] and all(ch in u["norm"] for ch in inits):
                        found_uid = u["id"]
                        break
            if found_uid:
                break
        if found_uid:
            p.user_id = found_uid
            db.add(p)
            linked += 1
    if linked:
        db.commit()
    return {"linked": int(linked), "scanned": len(pubs)}


# -----------------------------
# Department→Faculty mapping import (from Excel)
# -----------------------------

@router.post("/import/dept_map")
def import_dept_map(
    authorization: Optional[str] = Header(default=None),
    file: UploadFile = File(..., description="aku_faculties_departments.xlsx (sheet1: 'Факультет','Кафедра')"),
):
    """Upload Excel with columns 'Факультет' and 'Кафедра' on the first sheet and
    save department→faculty mapping to uploads/_dept_map.json used by statistics.
    """
    require_admin(authorization)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(UPLOAD_DIR, f"_dept_map_upload_{file.filename}")
    with open(temp_path, "wb") as out:
        out.write(file.file.read())

    try:
        # Try pandas, fallback to openpyxl
        mapping: dict[str, str] = {}
        try:
            import pandas as pd  # type: ignore
            df = pd.read_excel(temp_path, sheet_name=0)
            cols = {c.strip().lower(): c for c in df.columns if isinstance(c, str)}
            col_fac = cols.get('факультет')
            col_dep = cols.get('кафедра')
            if not (col_fac and col_dep):
                raise RuntimeError("Expected columns 'Факультет' and 'Кафедра' on sheet 1")
            for _, row in df.iterrows():
                fac = str(row[col_fac]).strip()
                dep = str(row[col_dep]).strip()
                if fac and dep and fac.lower() != 'nan' and dep.lower() != 'nan':
                    mapping[dep] = fac
        except Exception:
            # openpyxl minimal parser
            try:
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(temp_path, read_only=True)
                ws = wb.worksheets[0]
                # find headers row
                headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))[0].parent[1]]  # type: ignore
            except Exception:
                raise HTTPException(status_code=400, detail="Failed to read Excel. Ensure it has 'Факультет' and 'Кафедра'.")
        # Write mapping JSON
        if not mapping:
            raise HTTPException(status_code=400, detail="No mapping rows found. Check column names 'Факультет' and 'Кафедра'.")
        out_path = os.path.join(UPLOAD_DIR, "_dept_map.json")
        import json
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(mapping, f, ensure_ascii=False, indent=2)
        return {"status": "ok", "pairs": len(mapping), "path": "/uploads/_dept_map.json"}
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


@router.get("/maintenance/dept_map")
def get_dept_map(db: Session = Depends(get_db), _=Depends(require_admin)):
    """Return the active department→faculty mapping loaded by publications router."""
    try:
        from .routers_publications import _load_dept_map  # type: ignore
        m = _load_dept_map()
        return {"count": len(m), "mapping": m}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot load mapping: {e}")


@router.post("/import/dept_map_json")
def import_dept_map_json(
    payload: dict[str, str] = Body(..., description="JSON object: { 'Кафедра': 'Факультет', ... }"),
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    """Accept department→faculty mapping as JSON and save to uploads/_dept_map.json.
    Keys are department names, values are faculty names.
    """
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    if not isinstance(payload, dict) or not payload:
        raise HTTPException(status_code=400, detail="Body must be a non-empty JSON object")
    out_path = os.path.join(UPLOAD_DIR, "_dept_map.json")
    import json
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({str(k): str(v) for k, v in payload.items()}, f, ensure_ascii=False, indent=2)
    return {"status": "ok", "pairs": len(payload), "path": "/uploads/_dept_map.json"}


@router.post("/import/faculties_departments_fio")
def import_faculties_departments_fio(
    authorization: Optional[str] = Header(default=None),
    file: UploadFile = File(..., description="faculties_departments_fio.xlsx (columns: ФИО, Кафедра, Факультет) on sheet 1"),
    db: Session = Depends(get_db),
):
    """Import Excel with columns ФИО, Кафедра, Факультет.
    For each row:
      - Find existing User by exact normalized full name and update department/faculty.
      - If not found, CREATE a new User with generated unique login, role='teacher', active=1,
        and set department/faculty from the row (created_source='import').
      - Add/merge department→faculty to uploads/_dept_map.json.
    Returns created/updated counters and few examples.
    """
    require_admin(authorization)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(UPLOAD_DIR, f"_fio_map_{file.filename}")
    with open(temp_path, "wb") as out:
        out.write(file.file.read())

    # helpers
    def _norm_name(s: str) -> str:
        return (
            (s or "")
            .lower()
            .replace("\u00A0", " ")
            .replace(" ", "")
            .replace(".", "")
            .replace(",", "")
        )

    try:
        # Load Excel
        try:
            import pandas as pd  # type: ignore
            df = pd.read_excel(temp_path, sheet_name=0)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Cannot read Excel: {e}")
        # Resolve columns by case-insensitive match
        cols = {str(c).strip().lower(): c for c in df.columns}
        col_fio = cols.get('фио') or cols.get('ф.и.о') or cols.get('фио авторы')
        col_dep = cols.get('кафедра')
        col_fac = cols.get('факультет')
        if not (col_fio and col_dep and col_fac):
            raise HTTPException(status_code=400, detail="Expected columns: 'ФИО', 'Кафедра', 'Факультет' on sheet 1")

        # Build name->(dep,fac)
        items: list[tuple[str, str, str]] = []
        for _, row in df.iterrows():
            fio = str(row[col_fio]).strip()
            dep = str(row[col_dep]).strip()
            fac = str(row[col_fac]).strip()
            if fio and dep and fac and fio.lower() != 'nan' and dep.lower() != 'nan' and fac.lower() != 'nan':
                items.append((fio, dep, fac))

        if not items:
            raise HTTPException(status_code=400, detail="No data rows found in Excel")

        # Load and merge existing mapping JSON
        import json
        dept_map_path = os.path.join(UPLOAD_DIR, "_dept_map.json")
        dept_map: dict[str, str] = {}
        if os.path.isfile(dept_map_path):
            try:
                with open(dept_map_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    dept_map = {str(k): str(v) for k, v in data.items()}
            except Exception:
                dept_map = {}

        # Index users by normalized name and by fuzzy keys: last + first initial
        users = db.execute(select(User)).scalars().all()
        u_by_norm: dict[str, User] = {}
        fuzzy_index: dict[tuple[str, str], list[User]] = {}
        for u in users:
            nm = (u.full_name or '').replace('\u00A0', ' ').strip()
            if not nm:
                continue
            u_by_norm[_norm_name(nm)] = u
            parts = [p for p in nm.split() if p]
            if not parts:
                continue
            last = parts[0].lower()
            first_init = parts[1][0].lower() if len(parts) > 1 and parts[1] else ''
            key = (last, first_init)
            fuzzy_index.setdefault(key, []).append(u)

        created_users = 0
        updated_users = 0
        merged_pairs = 0
        examples: list[dict] = []
        for fio, dep, fac in items:
            # Merge mapping
            if dept_map.get(dep) != fac:
                dept_map[dep] = fac
                merged_pairs += 1
            # Update user if matches by full name
            key = _norm_name(fio)
            u = u_by_norm.get(key)
            # Fuzzy fallback: last name + first initial unique match
            if u is None:
                fio_clean = fio.replace('\u00A0', ' ').strip()
                parts = [p for p in fio_clean.split() if p]
                if parts:
                    last = parts[0].lower()
                    first_init = parts[1][0].lower() if len(parts) > 1 and parts[1] else ''
                    cand_list = fuzzy_index.get((last, first_init), [])
                    if len(cand_list) == 1:
                        u = cand_list[0]
            if u is not None:
                changed = False
                if (u.department or '').strip() != dep:
                    u.department = dep
                    changed = True
                if (u.faculty or '').strip() != fac:
                    u.faculty = fac
                    changed = True
                if changed:
                    db.add(u)
                    updated_users += 1
            else:
                # Create a new user
                from .models import User as _User
                from sqlalchemy import select as _select
                s = get_settings()
                # Generate a simple unique login from full name
                base = (
                    fio.replace("\u00A0", " ")
                    .lower()
                    .replace(" ", "")
                    .replace(",", "")
                    .replace(".", "")
                ) or "user"
                login = base
                # ensure uniqueness
                n = 0
                while db.execute(_select(_User).where(_User.login == login)).scalar_one_or_none() is not None:
                    n += 1
                    login = f"{base}{n}"
                # Setup initial password for compatibility (user can change later)
                import hashlib
                pw_plain = "123456"
                pw_hash = hashlib.sha256((s.PASSWORD_SALT + pw_plain).encode("utf-8")).hexdigest()
                new_user = _User(
                    login=login,
                    full_name=fio,
                    role="teacher",
                    faculty=fac,
                    department=dep,
                    position="",
                    degree="",
                    active=1,
                    password_hash=pw_hash,
                    initial_password=pw_plain,
                    created_source='import',
                )
                db.add(new_user)
                created_users += 1
                if len(examples) < 5:
                    examples.append({"user_id": new_user.id if getattr(new_user, 'id', None) else -1, "full_name": fio, "department": dep, "faculty": fac})

        if updated_users:
            db.commit()
        # Write back mapping JSON
        with open(dept_map_path, "w", encoding="utf-8") as f:
            json.dump(dept_map, f, ensure_ascii=False, indent=2)

        return {"status": "ok", "users_created": int(created_users), "users_updated": int(updated_users), "pairs_merged": int(merged_pairs), "examples": examples[:5]}
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass
