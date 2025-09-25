from __future__ import annotations
from typing import List, Optional
from fastapi import APIRouter, Depends, Query, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select, func, or_, and_, desc, asc
import io

try:
    from openpyxl import Workbook  # type: ignore
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore

from .db import get_db
from .models import Publication, Author, Source, User
from .schemas import SearchResponse, PageMeta, PublicationOut
from .config import get_settings

router = APIRouter(prefix="/search", tags=["search"])


def _apply_common_filters(stmt, filters, joins, need_join_authors, need_join_source):
    if need_join_authors:
        stmt = stmt.join(Publication.authors)
    if need_join_source:
        stmt = stmt.join(Publication.source)
    if filters:
        stmt = stmt.where(and_(*filters))
    return stmt


@router.get("/stats")
def search_stats(
    q: Optional[str] = Query(default=None),
    year_min: Optional[int] = Query(default=None),
    year_max: Optional[int] = Query(default=None),
    quartiles: Optional[List[str]] = Query(default=None),
    authors: Optional[List[int]] = Query(default=None),
    sources: Optional[List[int]] = Query(default=None),
    issn: Optional[str] = Query(default=None),
    source_type: Optional[str] = Query(default=None),
    upload_source: Optional[str] = Query(default=None, description="kokson|scopus|manual"),
    citations_min: Optional[int] = Query(default=None),
    citations_max: Optional[int] = Query(default=None),
    percentile_min: Optional[int] = Query(default=None),
    percentile_max: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Return aggregate stats for KPI and charts used on stats page.
    Applies the same filter semantics as /search (only approved publications).
    """
    filters = [Publication.status == "approved"]
    join_authors = False
    join_source = False

    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)
    if quartiles:
        filters.append(Publication.quartile.in_(quartiles))
    if sources:
        filters.append(Publication.source_id.in_(sources))
    if source_type:
        join_source = True
        filters.append(Source.type == source_type)
    if upload_source:
        if upload_source == 'scopus':
            filters.append(or_(Publication.upload_source == 'scopus', Publication.scopus_url.is_not(None)))
        else:
            filters.append(Publication.upload_source == upload_source)
    if issn:
        join_source = True
        filters.append(Source.issn.ilike(f"%{issn.strip()}%"))
    if citations_min is not None:
        filters.append(Publication.citations_count >= citations_min)
    if citations_max is not None:
        filters.append(Publication.citations_count <= citations_max)
    if percentile_min is not None:
        filters.append(Publication.percentile_2024 >= percentile_min)
    if percentile_max is not None:
        filters.append(Publication.percentile_2024 <= percentile_max)
    if q:
        join_authors = True
        join_source = True
        q_like = f"%{q.strip()}%"
        filters.append(or_(
            Publication.title.ilike(q_like),
            Publication.doi.ilike(q_like),
            Source.name.ilike(q_like),
            Author.display_name.ilike(q_like),
        ))
    if authors:
        join_authors = True
        filters.append(Author.id.in_(authors))

    # Helper to start a select with joins
    def make_base_select(cols):
        stmt = select(*cols)
        if join_authors:
            stmt = stmt.select_from(Publication).join(Publication.authors)
        else:
            stmt = stmt.select_from(Publication)
        if join_source:
            stmt = stmt.join(Publication.source)
        if filters:
            stmt = stmt.where(and_(*filters))
        return stmt

    # KPI
    total_pubs = db.execute(make_base_select([func.count(func.distinct(Publication.id))])).scalar() or 0
    # distinct authors over filtered publications
    stmt_auth_cnt = make_base_select([func.count(func.distinct(Author.id))])
    if not join_authors:
        stmt_auth_cnt = stmt_auth_cnt.join(Publication.authors)
    total_authors = db.execute(stmt_auth_cnt).scalar() or 0
    # distinct sources
    stmt_src_cnt = make_base_select([func.count(func.distinct(Source.id))])
    if not join_source:
        stmt_src_cnt = stmt_src_cnt.join(Publication.source)
    total_sources = db.execute(stmt_src_cnt).scalar() or 0
    avg_per_author = float(total_pubs) / float(total_authors or 1)

    # Yearly: publications and citations sum per year
    yearly_rows = db.execute(make_base_select([Publication.year, func.count(func.distinct(Publication.id)), func.sum(Publication.citations_count)])
                            .group_by(Publication.year)
                            .order_by(Publication.year)).all()
    yearly = [{"year": int(y), "publications": int(cnt or 0), "citations": int(cits or 0)} for (y, cnt, cits) in yearly_rows]

    # Top authors
    stmt_top_auth = make_base_select([Author.display_name, func.count(func.distinct(Publication.id))])
    if not join_authors:
        stmt_top_auth = stmt_top_auth.join(Publication.authors)
    top_authors_rows = db.execute(stmt_top_auth.group_by(Author.display_name).order_by(desc(func.count(func.distinct(Publication.id)))).limit(20)).all()
    top_authors = [{"author": n, "count": int(c or 0)} for (n, c) in top_authors_rows]

    # Top sources
    stmt_top_src = make_base_select([Source.name, func.count(func.distinct(Publication.id))])
    if not join_source:
        stmt_top_src = stmt_top_src.join(Publication.source)
    top_sources_rows = db.execute(stmt_top_src.group_by(Source.name).order_by(desc(func.count(func.distinct(Publication.id)))).limit(20)).all()
    top_sources = [{"source": n or "", "count": int(c or 0)} for (n, c) in top_sources_rows]

    # Quartiles distribution
    quart_rows = db.execute(make_base_select([Publication.quartile, func.count(func.distinct(Publication.id))])
                           .where(Publication.quartile.is_not(None))
                           .group_by(Publication.quartile)
                           .order_by(Publication.quartile)).all()
    quartiles_out = [{"quartile": qv or "", "count": int(c or 0)} for (qv, c) in quart_rows]

    return {
        "kpi": {
            "publications": int(total_pubs),
            "authors": int(total_authors),
            "sources": int(total_sources),
            "avg_per_author": round(avg_per_author, 2),
        },
        "yearly": yearly,
        "top_authors": top_authors,
        "top_sources": top_sources,
        "quartiles": quartiles_out,
    }


@router.get("", response_model=SearchResponse)
def search_publications(
    q: Optional[str] = Query(default=None, description="Keyword query across title/doi/author/source"),
    year_min: Optional[int] = Query(default=None),
    year_max: Optional[int] = Query(default=None),
    quartiles: Optional[List[str]] = Query(default=None),  # e.g., Q1,Q2
    authors: Optional[List[int]] = Query(default=None),
    sources: Optional[List[int]] = Query(default=None),
    issn: Optional[str] = Query(default=None),
    source_type: Optional[str] = Query(default=None, description="journal|conference"),
    upload_source: Optional[str] = Query(default=None, description="kokson|scopus|manual"),
    citations_min: Optional[int] = Query(default=None),
    citations_max: Optional[int] = Query(default=None),
    percentile_min: Optional[int] = Query(default=None),
    percentile_max: Optional[int] = Query(default=None),
    sort: str = Query(default="year_desc", description="year_desc|year_asc|citations_desc|citations_asc|title_asc|title_desc"),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=None),
    db: Session = Depends(get_db),
):
    settings = get_settings()
    if per_page is None:
        per_page = settings.PAGE_SIZE_DEFAULT
    per_page = max(1, min(per_page, settings.PAGE_SIZE_MAX))

    filters = [Publication.status == "approved"]
    joins = []

    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)

    if quartiles:
        filters.append(Publication.quartile.in_(quartiles))

    if sources:
        filters.append(Publication.source_id.in_(sources))

    if source_type:
        joins.append("source")
        filters.append(Source.type == source_type)
    if upload_source:
        # restrict by upload_source
        if upload_source == 'scopus':
            # Include legacy scopus imports that may lack the explicit flag
            filters.append(or_(Publication.upload_source == 'scopus', Publication.scopus_url.is_not(None)))
        else:
            filters.append(Publication.upload_source == upload_source)

    if issn:
        joins.append("source")
        filters.append(Source.issn.ilike(f"%{issn.strip()}%"))
    if upload_source:
        # Align semantics with /search and /search/stats
        if upload_source == 'scopus':
            filters.append(or_(Publication.upload_source == 'scopus', Publication.scopus_url.is_not(None)))
        else:
            filters.append(Publication.upload_source == upload_source)

    if citations_min is not None:
        filters.append(Publication.citations_count >= citations_min)
    if citations_max is not None:
        filters.append(Publication.citations_count <= citations_max)

    if percentile_min is not None:
        filters.append(Publication.percentile_2024 >= percentile_min)
    if percentile_max is not None:
        filters.append(Publication.percentile_2024 <= percentile_max)

    # text query across title, doi, author name, source name
    if q:
        q_like = f"%{q.strip()}%"
        # Left join authors and source for filtering
        joins.append("authors")
        joins.append("source")
        filters.append(
            or_(
                Publication.title.ilike(q_like),
                Publication.doi.ilike(q_like),
                Source.name.ilike(q_like),
                Author.display_name.ilike(q_like),
            )
        )

    if authors:
        joins.append("authors")
        filters.append(Author.id.in_(authors))

    stmt = select(Publication).options(
        joinedload(Publication.source),
        joinedload(Publication.authors),
    )

    # apply joins only once
    need_join_authors = "authors" in joins
    need_join_source = "source" in joins
    if need_join_authors:
        stmt = stmt.join(Publication.authors)
    if need_join_source:
        stmt = stmt.join(Publication.source)

    if filters:
        stmt = stmt.where(and_(*filters))

    # total count
    count_stmt = select(func.count(func.distinct(Publication.id)))
    if need_join_authors:
        count_stmt = count_stmt.join(Publication.authors)
    if need_join_source:
        count_stmt = count_stmt.join(Publication.source)
    if filters:
        count_stmt = count_stmt.where(and_(*filters))

    total: int = db.execute(count_stmt).scalar_one()

    # sorting
    if sort == "year_asc":
        stmt = stmt.order_by(asc(Publication.year), Publication.id)
    elif sort == "citations_desc":
        stmt = stmt.order_by(desc(Publication.citations_count), desc(Publication.year), Publication.id)
    elif sort == "citations_asc":
        stmt = stmt.order_by(asc(Publication.citations_count), desc(Publication.year), Publication.id)
    elif sort == "title_asc":
        stmt = stmt.order_by(asc(Publication.title), Publication.id)
    elif sort == "title_desc":
        stmt = stmt.order_by(desc(Publication.title), Publication.id)
    else:
        stmt = stmt.order_by(desc(Publication.year), Publication.id)

    # pagination
    offset = (page - 1) * per_page
    stmt = stmt.offset(offset).limit(per_page)

    result = db.execute(stmt).scalars().unique().all()

    total_pages = (total + per_page - 1) // per_page if total else 1

    return SearchResponse(
        meta=PageMeta(page=page, per_page=per_page, total=total, total_pages=total_pages),
        items=[PublicationOut.model_validate(pub) for pub in result],
    )


@router.get("/mine", response_model=SearchResponse)
def my_publications(
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=None),
    db: Session = Depends(get_db),
    x_user_id: int | None = Header(default=None, alias="X-User-Id"),
):
    """Return publications for the currently logged-in user by linking through Author.user_id.
    Frontend provides X-User-Id; we filter approved publications where any author has user_id == X-User-Id.
    """
    settings = get_settings()
    if per_page is None:
        per_page = settings.PAGE_SIZE_DEFAULT
    per_page = max(1, min(per_page, settings.PAGE_SIZE_MAX))
    if not x_user_id:
        return SearchResponse(meta=PageMeta(page=1, per_page=per_page, total=0, total_pages=1), items=[])

    # Base query: approved publications joined with authors filtered by user_id
    base = (
        select(Publication)
        .join(Publication.authors)
        .where(Publication.status == "approved", Author.user_id == x_user_id)
    )

    # Count distinct publications
    total = db.execute(select(func.count(func.distinct(Publication.id))).select_from(base.subquery())).scalar() or 0

    # Page
    offset = (page - 1) * per_page
    items = db.execute(base.order_by(desc(Publication.year), Publication.id).offset(offset).limit(per_page)).scalars().unique().all()
    total_pages = (total + per_page - 1) // per_page if total else 1
    return SearchResponse(
        meta=PageMeta(page=page, per_page=per_page, total=int(total), total_pages=total_pages),
        items=[PublicationOut.model_validate(p) for p in items],
    )


def _name_variants(full_name: str) -> list[str]:
    parts = [p for p in (full_name or "").replace("\u00A0", " ").strip().split() if p]
    if not parts:
        return []
    last = parts[0]
    initials_compact = "".join([f"{p[0]}." for p in parts[1:]])  # S.S.
    initials_spaced = " ".join([f"{p[0]}." for p in parts[1:]])  # S. S.
    c: set[str] = set()
    c.add(full_name)
    if initials_compact:
        c.add(f"{last} {initials_compact}")
        c.add(f"{initials_compact} {last}")
    if initials_spaced:
        c.add(f"{last} {initials_spaced}")
        c.add(f"{initials_spaced} {last}")
    return sorted({x.strip() for x in c if x.strip()})

def _translit_ru_to_en(s: str) -> str:
    table = {
        "А":"A","Б":"B","В":"V","Г":"G","Д":"D","Е":"E","Ё":"E","Ж":"Zh","З":"Z","И":"I","Й":"Y","К":"K","Л":"L","М":"M","Н":"N","О":"O","П":"P","Р":"R","С":"S","Т":"T","У":"U","Ф":"F","Х":"Kh","Ц":"Ts","Ч":"Ch","Ш":"Sh","Щ":"Sch","Ы":"Y","Э":"E","Ю":"Yu","Я":"Ya",
        "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"zh","з":"z","и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o","п":"p","р":"r","с":"s","т":"t","у":"u","ф":"f","х":"kh","ц":"ts","ч":"ch","ш":"sh","щ":"sch","ы":"y","э":"e","ю":"yu","я":"ya",
    }
    return "".join([table.get(ch, ch) for ch in s])

def _norm(s: str) -> str:
    return (
        (s or "")
        .lower()
        .replace("\u00a0", " ")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", "")
        .replace("-", "")
    )

def _extract_author_last_and_inits(author_display: str) -> tuple[str, str]:
    """Extract normalized last name token and initials from an author's display string.
    Example: 'A.Z., Bekeshev, A. Z.' -> ('bekeshev', 'az')
    """
    import re
    txt = (author_display or "").replace("\u00A0", " ")
    # initials like A.Z. or A. Z.
    inits = ''.join(re.findall(r"([A-Za-zА-Яа-яЁё])\.", txt)).lower()
    # tokens consisting of letters (ignore punctuation)
    tokens = re.findall(r"[A-Za-zА-Яа-яЁёІіӘәҒғҚқҢңӨөҰұҮүҺһ]+", txt)
    # choose the longest token length>=3 as last name candidate
    last_token = ''
    for t in tokens:
        if len(t) >= 3:
            last_token = t
            break
    return (_norm(last_token), inits)


@router.get("/authors/{author_id}")
def author_detail(author_id: int, db: Session = Depends(get_db)):
    """Return author card with optional matched user profile and all publications."""
    # Load author and publications
    a: Optional[Author] = db.get(Author, author_id)
    if not a:
        return {"error": "Author not found"}

    pubs = db.execute(
        select(Publication)
        .join(Publication.authors)
        .options(joinedload(Publication.source), joinedload(Publication.authors))
        .where(Author.id == author_id, Publication.status == "approved")
        .order_by(desc(Publication.year), Publication.id)
    ).scalars().unique().all()

    # Try to match User by equality, last name, or name_variants
    candidate: User | None = db.execute(
        select(User).where(func.lower(User.full_name) == func.lower(a.display_name))
    ).scalars().first()

    # 2) shortlist by detected last name + initials
    if not candidate:
        an_last, an_inits = _extract_author_last_and_inits(a.display_name)
        if an_last:
            users_all: list[User] = db.execute(select(User).limit(5000)).scalars().all()
            shortlist2: list[User] = []
            for u in users_all:
                parts_u = [p for p in (u.full_name or "").replace("\u00A0", " ").split() if p]
                if not parts_u:
                    continue
                last_ru = parts_u[0]
                last_en = _translit_ru_to_en(last_ru)
                if _norm(last_en) == an_last or _norm(last_ru) == an_last:
                    # check initials if present
                    u_inits_cyr = ''.join([p[0] for p in parts_u[1:]]).lower()
                    # transliterate initials to latin for fair compare
                    u_inits_lat = ''.join([_translit_ru_to_en(ch)[:1].lower() for ch in u_inits_cyr])
                    need = set(an_inits)
                    if an_inits and not need.issubset(set(u_inits_lat)):
                        continue
                    shortlist2.append(u)
            if len(shortlist2) == 1:
                candidate = shortlist2[0]

    # 2b) fuzzy last-name matching (to tolerate translit variants: Bogatariev vs Botagariev)
    if not candidate:
        try:
            import difflib
        except Exception:
            difflib = None  # type: ignore
        if difflib is not None:
            an_last, an_inits = _extract_author_last_and_inits(a.display_name)
            if an_last:
                users_all: list[User] = db.execute(select(User).limit(5000)).scalars().all()
                best_u: User | None = None
                best_score = 0.0
                for u in users_all:
                    parts_u = [p for p in (u.full_name or "").replace("\u00A0", " ").split() if p]
                    if not parts_u:
                        continue
                    last_ru = parts_u[0]
                    last_en = _translit_ru_to_en(last_ru)
                    score = difflib.SequenceMatcher(None, _norm(last_en), an_last).ratio()
                    if score >= 0.82:  # tolerant threshold for 1-2 letter differences
                        # Check initials if present
                        u_inits_cyr = ''.join([p[0] for p in parts_u[1:]]).lower()
                        u_inits_lat = ''.join([_translit_ru_to_en(ch)[:1].lower() for ch in u_inits_cyr])
                        need = set(an_inits)
                        if an_inits and not need.issubset(set(u_inits_lat)):
                            continue
                        if score > best_score:
                            best_score = score
                            best_u = u
                if best_u is not None:
                    candidate = best_u

    # 3) broader fallback: try match against ALL users by variants (handles cases where author starts with initials, e.g. "A.Z., Bekeshev, A. Z.")
    if not candidate:
        an_norm = _norm(a.display_name)
        users_all: list[User] = db.execute(select(User).limit(5000)).scalars().all()
        for u in users_all:
            variants: list[str] = []
            try:
                if u.name_variants:
                    import json
                    vv = json.loads(u.name_variants)
                    if isinstance(vv, list):
                        variants.extend([str(x) for x in vv])
            except Exception:
                pass
            # build initials and last name variants (ru/en)
            parts = [p for p in (u.full_name or "").replace("\u00A0", " ").split() if p]
            last_ru = parts[0] if parts else ""
            last_en = _translit_ru_to_en(last_ru)
            inits = [p[0] for p in parts[1:]]
            # add common patterns to variants
            if last_ru:
                variants += [f"{last_ru} " + ".".join(inits) + "."] if inits else []
                variants += [" ".join([f"{ch}." for ch in inits]) + f" {last_ru}"] if inits else []
            if last_en:
                variants += [f"{last_en} " + ".".join(inits) + "."] if inits else []
                variants += [" ".join([f"{ch}." for ch in inits]) + f" {last_en}"] if inits else []
            # final check against normalized author name
            vnorms = [_norm(v) for v in variants if v]
            if any(vn and vn in an_norm for vn in vnorms) or any(an_norm and an_norm in vn for vn in vnorms):
                candidate = u
                break

    user_info = None
    if candidate:
        user_info = {
            "id": candidate.id,
            "full_name": candidate.full_name,
            "faculty": candidate.faculty,
            "department": candidate.department,
            "position": candidate.position,
            "degree": candidate.degree,
        }

    return {
        "author": {"id": a.id, "display_name": a.display_name},
        "user": user_info,
        "publications": [PublicationOut.model_validate(p) for p in pubs],
    }


@router.get("/authors/{author_id}/export")
def author_export(
    author_id: int,
    fmt: str = Query(default="csv", description="csv|xlsx"),
    db: Session = Depends(get_db),
):
    """Export all approved publications of the author to CSV/XLSX."""
    a: Optional[Author] = db.get(Author, author_id)
    if not a:
        def iter_err():
            yield "Author not found"
        return StreamingResponse(iter_err(), media_type="text/plain")

    rows = db.execute(
        select(Publication)
        .join(Publication.authors)
        .options(joinedload(Publication.source), joinedload(Publication.authors))
        .where(Author.id == author_id, Publication.status == "approved")
        .order_by(desc(Publication.year), Publication.id)
    ).scalars().unique().all()

    if fmt.lower() == "csv":
        def iter_csv():
            header = [
                "id","year","title","authors","source","issn","doi","scopus_url","citations","quartile","percentile_2024","pdf_url"
            ]
            yield ",".join(header) + "\n"
            for p in rows:
                authors_str = "; ".join([au.display_name for au in p.authors])
                source_name = p.source.name if p.source else ""
                issn_val = p.source.issn if p.source else ""
                vals = [
                    str(p.id), str(p.year or ''), p.title.replace('"','""'), authors_str.replace('"','""'), source_name.replace('"','""'),
                    str(issn_val or ''), str(p.doi or ''), str(p.scopus_url or ''), str(p.citations_count or 0), str(p.quartile or ''), str(p.percentile_2024 or ''), str(p.pdf_url or ''),
                ]
                def q(v: str) -> str:
                    return f'"{v}"' if ("," in v or "\n" in v or '"' in v) else v
                yield ",".join([q(v) for v in vals]) + "\n"
        filename = f"author_{author_id}_export.csv"
        return StreamingResponse(iter_csv(), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})

    if fmt.lower() == "xlsx":
        if Workbook is None:
            def iter_err():
                yield "XLSX export not available"
            return StreamingResponse(iter_err(), media_type="text/plain")
        wb = Workbook()
        ws = wb.active
        ws.title = "Author"
        header = ["id","year","title","authors","source","issn","doi","scopus_url","citations","quartile","percentile_2024","pdf_url"]
        ws.append(header)
        for p in rows:
            authors_str = "; ".join([au.display_name for au in p.authors])
            source_name = p.source.name if p.source else ""
            issn_val = p.source.issn if p.source else ""
            ws.append([
                p.id, p.year or '', p.title, authors_str, source_name,
                issn_val or '', p.doi or '', p.scopus_url or '', p.citations_count or 0, p.quartile or '', p.percentile_2024 or '', p.pdf_url or ''
            ])
        # Autosize columns
        for col in ws.columns:
            max_len = 10
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"author_{author_id}_export.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

    def iter_fallback():
        yield "Unsupported format"
    return StreamingResponse(iter_fallback(), media_type="text/plain")


@router.get("/facets/faculties")
def facets_faculties(limit: int = 100, db: Session = Depends(get_db)):
    """Return list of faculties from Users with de-duplicated counts per faculty.
    Count uses DISTINCT normalized full_name to avoid duplicates from import.
    """
    def norm_expr(col):
        return func.lower(
            func.replace(
                func.replace(
                    func.replace(col, "\u00A0", " "),
                    ",", ""
                ),
                ".", ""
            )
        )
    rows = db.execute(
        select(
            User.faculty,
            func.count(func.distinct(norm_expr(User.full_name)))
        )
        .where(
            User.faculty != "",
            User.faculty.is_not(None),
            func.lower(User.faculty) != "nan",
        )
        .group_by(User.faculty)
        .order_by(desc(func.count(func.distinct(norm_expr(User.full_name)))))
        .limit(limit)
    ).all()
    return [{"name": r[0], "count": int(r[1])} for r in rows]


@router.get("/facets/faculties_pubs")
def facets_faculties_pubs(limit: int = 100, db: Session = Depends(get_db)):
    """Return faculties with people_count and publications_count.
    We compute people_count via SQL, and publications_count in Python by
    mapping normalized author strings to faculties using users' name_variants.
    """
    # 1) People counts (ordering baseline)
    def norm_expr(col):
        return func.lower(
            func.replace(
                func.replace(
                    func.replace(col, "\u00A0", " "),
                    ",", ""
                ),
                ".", ""
            )
        )
    fac_rows = db.execute(
        select(
            User.faculty,
            func.count(func.distinct(norm_expr(User.full_name))).label("people")
        )
        .where(
            User.faculty != "",
            User.faculty.is_not(None),
            func.lower(User.faculty) != "nan",
        )
        .group_by(User.faculty)
        .order_by(desc("people"))
        .limit(limit)
    ).all()

    # 2) Build normalized variant -> set(faculty) index from Users
    import json, re
    def _n(s: str) -> str:
        return (s or "").replace("\u00A0"," ").lower().replace(" ","").replace(".","").replace(",","")

    variant_to_faculties: dict[str, set[str]] = {}
    users_all: list[User] = db.execute(select(User).where(User.faculty.is_not(None))).scalars().all()
    for u in users_all:
        if not u.faculty:
            continue
        fac = u.faculty
        variants: list[str] = [u.full_name or ""]
        try:
            if u.name_variants:
                vv = json.loads(u.name_variants)
                if isinstance(vv, list):
                    variants.extend([str(x) for x in vv])
        except Exception:
            pass
        # also add simple patterns: "Last F." and with comma removed
        parts = [p for p in (u.full_name or '').replace('\u00A0',' ').split() if p]
        if parts:
            last = parts[0]
            inits = [p[0] for p in parts[1:]]
            if inits:
                variants.append(f"{last} " + ".".join(inits) + ".")
                variants.append(", ".join([last, " ".join(parts[1:])]))
        for v in set(variants):
            key = _n(v)
            if not key:
                continue
            variant_to_faculties.setdefault(key, set()).add(fac)

    # 3) Scan approved publications once and accumulate matches per faculty
    fac_to_pubids: dict[str, set[int]] = {name: set() for name, _ in fac_rows}
    pubs = db.execute(
        select(Publication)
        .options(joinedload(Publication.authors))
        .where(Publication.status == "approved")
    ).scalars().unique().all()
    for p in pubs:
        pid = p.id
        for a in p.authors:
            # strip parentheses like (123456) and normalize
            a_raw = re.sub(r"\s*\([^)]*\)\s*", " ", (a.display_name or "")).strip()
            key = _n(a_raw)
            if key in variant_to_faculties:
                for fac in variant_to_faculties[key]:
                    if fac in fac_to_pubids:
                        fac_to_pubids[fac].add(pid)

    # 4) Build response
    out = []
    for name, people in fac_rows:
        pubs_cnt = len(fac_to_pubids.get(name, set()))
        out.append({"name": name, "people_count": int(people), "publications_count": pubs_cnt})
    return out


@router.get("/facets/departments_pubs")
def facets_departments_pubs(limit: int = 100, db: Session = Depends(get_db)):
    """Return departments with people_count and publications_count using Python-side variant mapping (fast enough for ~1k rows)."""
    def norm_expr(col):
        return func.lower(
            func.replace(
                func.replace(
                    func.replace(col, "\u00A0", " "),
                    ",", ""
                ),
                ".", ""
            )
        )
    import json, re
    def _n(s: str) -> str:
        return (s or "").replace("\u00A0"," ").lower().replace(" ","").replace(".","").replace(",","")

    dep_rows = db.execute(
        select(
            User.department,
            func.count(func.distinct(norm_expr(User.full_name))).label("people")
        )
        .where(
            User.department != "",
            User.department.is_not(None),
            func.lower(User.department) != "nan",
        )
        .group_by(User.department)
        .order_by(desc("people"))
        .limit(limit)
    ).all()

    # Build variant -> departments map
    variant_to_deps: dict[str, set[str]] = {}
    users_all: list[User] = db.execute(select(User).where(User.department.is_not(None))).scalars().all()
    for u in users_all:
        if not u.department:
            continue
        dep = u.department
        variants: list[str] = [u.full_name or ""]
        try:
            if u.name_variants:
                vv = json.loads(u.name_variants)
                if isinstance(vv, list):
                    variants.extend([str(x) for x in vv])
        except Exception:
            pass
        parts = [p for p in (u.full_name or '').replace('\u00A0',' ').split() if p]
        if parts:
            last = parts[0]
            inits = [p[0] for p in parts[1:]]
            if inits:
                variants.append(f"{last} " + ".".join(inits) + ".")
                variants.append(", ".join([last, " ".join(parts[1:])]))
        for v in set(variants):
            key = _n(v)
            if not key:
                continue
            variant_to_deps.setdefault(key, set()).add(dep)

    dep_to_pubids: dict[str, set[int]] = {name: set() for name, _ in dep_rows}
    pubs = db.execute(
        select(Publication)
        .options(joinedload(Publication.authors))
        .where(Publication.status == "approved")
    ).scalars().unique().all()
    for p in pubs:
        pid = p.id
        for a in p.authors:
            a_raw = re.sub(r"\s*\([^)]*\)\s*", " ", (a.display_name or "")).strip()
            key = _n(a_raw)
            if key in variant_to_deps:
                for dep in variant_to_deps[key]:
                    if dep in dep_to_pubids:
                        dep_to_pubids[dep].add(pid)

    out = []
    for name, people in dep_rows:
        pubs_cnt = len(dep_to_pubids.get(name, set()))
        out.append({"name": name, "people_count": int(people), "publications_count": pubs_cnt})
    return out

@router.get("/facets/departments")
def facets_departments(limit: int = 100, db: Session = Depends(get_db)):
    """Return list of departments from Users with de-duplicated counts per department."""
    def norm_expr(col):
        return func.lower(
            func.replace(
                func.replace(
                    func.replace(col, "\u00A0", " "),
                    ",", ""
                ),
                ".", ""
            )
        )
    rows = db.execute(
        select(
            User.department,
            func.count(func.distinct(norm_expr(User.full_name)))
        )
        .where(
            User.department != "",
            User.department.is_not(None),
            func.lower(User.department) != "nan",
        )
        .group_by(User.department)
        .order_by(desc(func.count(func.distinct(norm_expr(User.full_name)))))
        .limit(limit)
    ).all()
    return [{"name": r[0], "count": int(r[1])} for r in rows]

@router.get("/faculty/export")
def faculty_export(
    faculty: str = Query(..., description="Faculty or department text (partial is OK)"),
    match: str = Query(default="broad", description="exact|initials|broad"),
    scope: str = Query(default="auto", description="auto|faculty|department"),
    fmt: str = Query(default="xlsx", description="csv|xlsx"),
    db: Session = Depends(get_db),
):
    """Export all publications for users belonging to given faculty.
    Matching strategy mirrors admin users_publications with 'match' parameter.
    """
    # Helper to normalize both DB column and input (remove spaces, dots, quotes, guillemets)
    def norm_expr(col):
        return func.lower(
            func.replace(
                func.replace(
                    func.replace(
                        func.replace(
                            func.replace(
                                func.replace(col, "\u00A0", " "),
                                ".", ""
                            ),
                            ",", ""
                        ),
                        " ", ""
                    ),
                    "«", ""
                ),
                "»", ""
            )
        )

    fac_norm = (
        faculty.replace("\u00A0", " ")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", "")
        .replace("«", "")
        .replace("»", "")
        .lower()
    )

    # Fetch users by faculty/department with partial match on normalized text
    users: list[User] = []
    if scope in ("auto", "faculty"):
        users = db.execute(
            select(User).where(norm_expr(User.faculty).like(f"%{fac_norm}%"))
        ).scalars().all()
    if not users and scope in ("auto", "department"):
        users = db.execute(
            select(User).where(norm_expr(User.department).like(f"%{fac_norm}%"))
        ).scalars().all()

    # Fallbacks: strip service words like 'кафедра/кафедрасы' and retry; also try raw ilike
    if not users:
        base_raw = faculty.replace("«", "").replace("»", "").strip()
        stripped = base_raw
        for token in ["кафедрасы", "кафедра", "каф."]:
            stripped = stripped.replace(token, "").strip()
        stripped_norm = (
            stripped.replace("\u00A0", " ")
            .replace(" ", "")
            .replace(".", "")
            .replace(",", "")
            .lower()
        )
        if scope in ("auto", "faculty"):
            users = db.execute(
                select(User).where(
                    or_(
                        norm_expr(User.faculty).like(f"%{stripped_norm}%"),
                        User.faculty.ilike(f"%{stripped}%"),
                    )
                )
            ).scalars().all()
        if not users and scope in ("auto", "department"):
            users = db.execute(
                select(User).where(
                    or_(
                        norm_expr(User.department).like(f"%{stripped_norm}%"),
                        User.department.ilike(f"%{stripped}%"),
                    )
                )
            ).scalars().all()
    if not users:
        def iter_err():
            yield "No users for faculty/department"
        return StreamingResponse(iter_err(), media_type="text/plain")

    pub_ids: set[int] = set()
    for u in users:
        raw = (u.full_name or '').replace("\u00A0", " ").strip()
        norm = raw.lower().replace(".", "").replace(",", "").replace(" ", "")
        parts = [p for p in raw.split() if p]
        last = parts[0].lower() if parts else ''
        inits = ''.join([p[0].lower() for p in parts[1:]])

        # A) exact equality
        ids_eq = db.execute(
            select(Publication.id)
            .join(Publication.authors)
            .where(norm_expr(Author.display_name) == norm)
        ).all()

        pub_ids.update([row[0] for row in ids_eq])

        if match in ("initials", "broad") and last:
            # B) last + initials contained
            cond = norm_expr(Author.display_name).like(f"%{last}%")
            for ch in inits:
                cond = cond & norm_expr(Author.display_name).like(f"%{ch}%")
            ids_fb = db.execute(
                select(Publication.id).join(Publication.authors).where(cond)
            ).all()
            pub_ids.update([row[0] for row in ids_fb])

        if match == "broad":
            # C) name variants OR ilike
            variants = _name_variants(u.full_name)
            if variants:
                conds = [Author.display_name.ilike(f"%{v}%") for v in variants]
                ids_var = db.execute(
                    select(Publication.id).join(Publication.authors).where(or_(*conds))
                ).all()
                pub_ids.update([row[0] for row in ids_var])

    if not pub_ids:
        # Last-resort broad match: any publication whose author contains user's last name (normalized), without requiring initials
        last_tokens: set[str] = set()
        for u in users:
            parts = [p for p in (u.full_name or '').replace("\u00A0", " ").split() if p]
            if parts:
                last_tokens.add(
                    (parts[0].lower().replace(".", "").replace(",", "").replace(" ", ""))
                )
                # also add transliterated last name
                last_tokens.add(_translit_ru_to_en(parts[0]).lower().replace(" ", ""))
        cond = None
        for ln in sorted({t for t in last_tokens if t}):
            c = norm_expr(Author.display_name).like(f"%{ln}%")
            cond = c if cond is None else (cond | c)
        if cond is not None:
            ids_any = db.execute(
                select(Publication.id).join(Publication.authors).where(cond)
            ).all()
            pub_ids.update([row[0] for row in ids_any])

    if not pub_ids:
        # Return empty file with headers instead of error to simplify reporting
        if fmt.lower() == "csv":
            def iter_csv_empty():
                header = [
                    "id","year","title","authors","source","issn","doi","scopus_url",
                    "citations","quartile","percentile_2024","pdf_url","faculty","department"
                ]
                yield ",".join(header) + "\n"
            return StreamingResponse(
                iter_csv_empty(),
                media_type="text/csv",
                headers={
                    "Content-Disposition": "attachment; filename=otchet.csv; filename*=UTF-8''%D0%BE%D1%82%D1%87%D0%B5%D1%82.csv"
                },
            )
        # XLSX empty
        if Workbook is None:
            def iter_err():
                yield ""
            return StreamingResponse(iter_err(), media_type="text/plain")
        wb = Workbook()
        ws = wb.active
        ws.title = "Faculty"
        header = [
            "id","year","title","authors","source","issn","doi","scopus_url",
            "citations","quartile","percentile_2024","pdf_url","faculty","department"
        ]
        ws.append(header)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=otchet.xlsx; filename*=UTF-8''%D0%BE%D1%82%D1%87%D0%B5%D1%82.xlsx"
            },
        )

    rows = db.execute(
        select(Publication)
        .where(Publication.id.in_(list(pub_ids)), Publication.status == "approved")
        .options(joinedload(Publication.source), joinedload(Publication.authors))
        .order_by(desc(Publication.year), Publication.id)
    ).scalars().unique().all()

    # Helper: try to resolve author_name to a concrete User among the selected users list
    def resolve_user_for_author(author_name: str) -> Optional[User]:
        import re
        # remove parenthetical IDs like '(...6504403163)'
        an_raw = re.sub(r"\s*\([^)]*\)\s*", " ", (author_name or '').replace('\u00A0', ' ')).strip()
        an = _norm(an_raw)
        # Capture initials including Latin digraphs before a dot, e.g., 'Zh.', 'Sh.'
        raw_initial_tokens = [m for m in re.findall(r"([A-Za-zА-Яа-яЁё]{1,3})\.", an_raw)]
        # Normalize each to Latin uppercase form comparable with user's transliterated initials
        def _tok_to_lat(tok: str) -> str:
            # If Cyrillic single letter, transliterate (Ж->Zh, Ч->Ch, Ш->Sh)
            if any('А' <= ch <= 'я' or ch in 'ЁёІіӘәҒғҚқҢңӨөҰұҮүҺһ' for ch in tok):
                return _translit_ru_to_en(tok)[:2].upper()
            return tok[:2].upper()
        initials_in_author_lat = [_tok_to_lat(t) for t in raw_initial_tokens]

        best_u: Optional[User] = None
        best_score = 0.0
        for u in users:
            parts = [p for p in (u.full_name or '').replace('\u00A0', ' ').split() if p]
            if not parts:
                continue
            last_ru = parts[0]
            first_ru = parts[1] if len(parts) > 1 else ''
            last_en = _translit_ru_to_en(last_ru)
            first_en = _translit_ru_to_en(first_ru)

            # 1) Last name must match (ru or translit)
            last_ok = (_norm(last_ru) in an) or (_norm(last_en) in an) or any((_norm(v) in an) or (an in _norm(v)) for v in _name_variants(u.full_name))
            if not last_ok:
                continue

            # 2) If initials are present in author string, require they are subset of user's initials
            if initials_in_author_lat:
                # Build user's transliterated initials as 2-char tokens (to match Zh/Ch/Sh)
                user_initial_tokens = []
                for p in parts[1:]:
                    lat = _translit_ru_to_en(p[:1]).upper()
                    user_initial_tokens.append(lat[:2])
                if not set(initials_in_author_lat).issubset(set(user_initial_tokens)):
                    continue
                # Additionally, ensure the author's FIRST initial matches user's FIRST initial
                # e.g., 'Abzal K.' should not match a user 'Nazar K.'
                author_first_init = initials_in_author_lat[0] if initials_in_author_lat else ''
                user_first_init = (user_initial_tokens[0] if user_initial_tokens else '')
                if author_first_init and user_first_init and author_first_init != user_first_init:
                    continue
            else:
                # 3) No initials given: require evidence of first name token (ru or translit) OR a name variant match
                first_ok = False
                if first_ru:
                    if (_norm(first_ru) in an) or (_norm(first_en) in an):
                        first_ok = True
                if not first_ok:
                    # try variants (e.g., 'Last F.' or 'F. Last')
                    variants = _name_variants(u.full_name)
                    vnorms = [_norm(v) for v in variants if v]
                    if any(vn and vn in an for vn in vnorms) or any(an and an in vn for vn in vnorms):
                        first_ok = True
                if not first_ok:
                    continue

            # score by length of last name match
            score = max(len(last_ru), len(last_en))
            if score > best_score:
                best_score = score
                best_u = u
        return best_u

    if fmt.lower() == "csv":
        def iter_csv():
            header = [
                "year","title","authors","source","issn","doi","scopus_url",
                "citations","quartile","percentile_2024","pdf_url","faculty","department"
            ]
            yield ",".join(header) + "\n"
            for p in rows:
                filtered_authors = [
                    a.display_name for a in p.authors
                    if resolve_user_for_author(a.display_name) is not None
                ]
                if not filtered_authors:
                    # Skip publications without authors from the selected faculty/department
                    continue
                authors_str = "; ".join(filtered_authors)
                source_name = p.source.name if p.source else ""
                issn_val = p.source.issn if p.source else ""
                vals = [
                    str(p.year or ''), p.title.replace('"','""'), authors_str.replace('"','""'), source_name.replace('"','""'),
                    str(issn_val or ''), str(p.doi or ''), str(p.scopus_url or ''), str(p.citations_count or 0), str(p.quartile or ''), str(p.percentile_2024 or ''), str(p.pdf_url or ''),
                    faculty,
                    faculty if scope == "department" else "",
                ]
                def q(v: str) -> str:
                    return f'"{v}"' if ("," in v or "\n" in v or '"' in v) else v
                yield ",".join([q(v) for v in vals]) + "\n"
        return StreamingResponse(
            iter_csv(),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=otchet.csv; filename*=UTF-8''%D0%BE%D1%82%D1%87%D0%B5%D1%82.csv"
            },
        )

    # default XLSX
    if Workbook is None:
        def iter_err():
            yield "XLSX not available"
        return StreamingResponse(iter_err(), media_type="text/plain")
    wb = Workbook()
    ws = wb.active
    ws.title = faculty[:31] or "Faculty"
    header = [
        "year","title","authors","source","issn","doi","scopus_url",
        "citations","quartile","percentile_2024","pdf_url","faculty","department"
    ]
    ws.append(header)
    for p in rows:
        filtered_authors = [
            a.display_name for a in p.authors
            if resolve_user_for_author(a.display_name) is not None
        ]
        if not filtered_authors:
            # Skip publications without authors from the selected faculty/department
            continue
        authors_str = "; ".join(filtered_authors)
        source_name = p.source.name if p.source else ""
        issn_val = p.source.issn if p.source else ""
        ws.append([
            p.year or '', p.title, authors_str, source_name,
            issn_val or '', p.doi or '', p.scopus_url or '', p.citations_count or 0, p.quartile or '', p.percentile_2024 or '', p.pdf_url or '',
            faculty,
            faculty if scope == "department" else "",
        ])
    for col in ws.columns:
        max_len = 10
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=otchet.xlsx; filename*=UTF-8''%D0%BE%D1%82%D1%87%D0%B5%D1%82.xlsx"
        },
    )


@router.get("/faculty/count")
def faculty_count(
    faculty: str = Query(..., description="Faculty or department text (partial is OK)"),
    match: str = Query(default="broad", description="exact|initials|broad"),
    scope: str = Query(default="auto", description="auto|faculty|department"),
    db: Session = Depends(get_db),
):
    """Return the number of publication rows that would be present in the export
    (i.e., publications having at least one author matched to the selected faculty/department).
    Reuses the same matching logic as faculty_export to ensure parity with CSV/XLSX.
    """
    # Reuse inner helpers from faculty_export
    def norm_expr(col):
        return func.lower(
            func.replace(
                func.replace(
                    func.replace(
                        func.replace(
                            func.replace(
                                func.replace(col, "\u00A0", " "),
                                ".", ""
                            ),
                            ",", ""
                        ),
                        " ", ""
                    ),
                    "«", ""
                ),
                "»", ""
            )
        )

    fac_norm = (
        faculty.replace("\u00A0", " ")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", "")
        .replace("«", "")
        .replace("»", "")
        .lower()
    )

    users: list[User] = []
    if scope in ("auto", "faculty"):
        users = db.execute(select(User).where(norm_expr(User.faculty).like(f"%{fac_norm}%"))).scalars().all()
    if not users and scope in ("auto", "department"):
        users = db.execute(select(User).where(norm_expr(User.department).like(f"%{fac_norm}%"))).scalars().all()
    if not users:
        base_raw = faculty.replace("«", "").replace("»", "").strip()
        stripped = base_raw
        for token in ["кафедрасы", "кафедра", "каф."]:
            stripped = stripped.replace(token, "").strip()
        stripped_norm = (
            stripped.replace("\u00A0", " ")
            .replace(" ", "")
            .replace(".", "")
            .replace(",", "")
            .lower()
        )
        if scope in ("auto", "faculty"):
            users = db.execute(
                select(User).where(
                    or_(
                        norm_expr(User.faculty).like(f"%{stripped_norm}%"),
                        User.faculty.ilike(f"%{stripped}%"),
                    )
                )
            ).scalars().all()
        if not users and scope in ("auto", "department"):
            users = db.execute(
                select(User).where(
                    or_(
                        norm_expr(User.department).like(f"%{stripped_norm}%"),
                        User.department.ilike(f"%{stripped}%"),
                    )
                )
            ).scalars().all()

    if not users:
        return {"count": 0}

    # Iterate over approved publications and count those that would appear in export
    rows = db.execute(
        select(Publication)
        .where(Publication.status == "approved")
        .options(joinedload(Publication.source), joinedload(Publication.authors))
    ).scalars().unique().all()

    # Use the same author filtering as in export to decide which publications produce rows
    def resolve_user_for_author(author_name: str) -> Optional[User]:
        import re
        an_raw = re.sub(r"\s*\([^)]*\)\s*", " ", (author_name or '').replace('\u00A0', ' ')).strip()
        an = _norm(an_raw)
        # Extract initials with digraphs support
        raw_initial_tokens = [m for m in re.findall(r"([A-Za-zА-Яа-яЁё]{1,3})\.", an_raw)]
        def _tok_to_lat(tok: str) -> str:
            if any('А' <= ch <= 'я' or ch in 'ЁёІіӘәҒғҚқҢңӨөҰұҮүҺһ' for ch in tok):
                return _translit_ru_to_en(tok)[:2].upper()
            return tok[:2].upper()
        initials_in_author_lat = [_tok_to_lat(t) for t in raw_initial_tokens]

        best_u: Optional[User] = None
        best_score = 0.0
        for u in users:
            parts = [p for p in (u.full_name or '').replace('\u00A0', ' ').split() if p]
            if not parts:
                continue
            last_ru = parts[0]
            first_ru = parts[1] if len(parts) > 1 else ''
            last_en = _translit_ru_to_en(last_ru)
            first_en = _translit_ru_to_en(first_ru)

            last_ok = (_norm(last_ru) in an) or (_norm(last_en) in an) or any((_norm(v) in an) or (an in _norm(v)) for v in _name_variants(u.full_name))
            if not last_ok:
                continue
            if initials_in_author_lat:
                user_initial_tokens = []
                for p in parts[1:]:
                    lat = _translit_ru_to_en(p[:1]).upper()
                    user_initial_tokens.append(lat[:2])
                if not set(initials_in_author_lat).issubset(set(user_initial_tokens)):
                    continue
                author_first_init = initials_in_author_lat[0] if initials_in_author_lat else ''
                user_first_init = (user_initial_tokens[0] if user_initial_tokens else '')
                if author_first_init and user_first_init and author_first_init != user_first_init:
                    continue
            else:
                first_ok = False
                if first_ru:
                    if (_norm(first_ru) in an) or (_norm(first_en) in an):
                        first_ok = True
                if not first_ok:
                    variants = _name_variants(u.full_name)
                    vnorms = [_norm(v) for v in variants if v]
                    if any(vn and vn in an for vn in vnorms) or any(an and an in vn for vn in vnorms):
                        first_ok = True
                if not first_ok:
                    continue
            score = max(len(last_ru), len(last_en))
            if score > best_score:
                best_score = score
                best_u = u
        return best_u

    count_rows = 0
    for p in rows:
        filtered_authors = [a.display_name for a in p.authors if resolve_user_for_author(a.display_name) is not None]
        if filtered_authors:
            count_rows += 1
    return {"count": count_rows}


@router.get("/facets/authors")
def facets_authors(
    q: Optional[str] = Query(default=None),
    year_min: Optional[int] = None,
    year_max: Optional[int] = None,
    quartiles: Optional[List[str]] = None,
    sources: Optional[List[int]] = None,
    issn: Optional[str] = None,
    source_type: Optional[str] = None,
    upload_source: Optional[str] = Query(default=None, description="kokson|scopus|manual"),
    citations_min: Optional[int] = None,
    citations_max: Optional[int] = None,
    percentile_min: Optional[int] = None,
    percentile_max: Optional[int] = None,
    query: Optional[str] = Query(default=None, description="filter authors by name"),
    limit: int = 50,
    db: Session = Depends(get_db),
):
    filters = [Publication.status == "approved"]
    joins: List[str] = ["authors"]
    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)
    if quartiles:
        filters.append(Publication.quartile.in_(quartiles))
    if sources:
        filters.append(Publication.source_id.in_(sources))
    if source_type:
        joins.append("source")
        filters.append(Source.type == source_type)
    if issn:
        joins.append("source")
        filters.append(Source.issn.ilike(f"%{issn.strip()}%"))
    if citations_min is not None:
        filters.append(Publication.citations_count >= citations_min)
    if citations_max is not None:
        filters.append(Publication.citations_count <= citations_max)
    if percentile_min is not None:
        filters.append(Publication.percentile_2024 >= percentile_min)
    if percentile_max is not None:
        filters.append(Publication.percentile_2024 <= percentile_max)
    if q:
        q_like = f"%{q.strip()}%"
        joins.append("source")
        filters.append(or_(Publication.title.ilike(q_like), Publication.doi.ilike(q_like), Source.name.ilike(q_like)))

    need_join_authors = True
    need_join_source = "source" in joins
    stmt = select(Author.id, Author.display_name, func.count(func.distinct(Publication.id))).join(Publication.authors)
    if need_join_source:
        stmt = stmt.join(Publication.source)
    if filters:
        stmt = stmt.where(and_(*filters))
    if query:
        stmt = stmt.where(Author.display_name.ilike(f"%{query.strip()}%"))
    stmt = stmt.group_by(Author.id, Author.display_name).order_by(desc(func.count(func.distinct(Publication.id)))).limit(limit)
    rows = db.execute(stmt).all()
    return [{"id": r[0], "name": r[1], "count": r[2]} for r in rows]


@router.get("/facets/sources")
def facets_sources(
    q: Optional[str] = Query(default=None),
    year_min: Optional[int] = None,
    year_max: Optional[int] = None,
    quartiles: Optional[List[str]] = None,
    authors: Optional[List[int]] = None,
    issn: Optional[str] = None,
    source_type: Optional[str] = None,
    citations_min: Optional[int] = None,
    citations_max: Optional[int] = None,
    percentile_min: Optional[int] = None,
    percentile_max: Optional[int] = None,
    query: Optional[str] = Query(default=None, description="filter sources by name"),
    limit: int = 50,
    db: Session = Depends(get_db),
):
    filters = [Publication.status == "approved"]
    joins: List[str] = ["source"]
    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)
    if quartiles:
        filters.append(Publication.quartile.in_(quartiles))
    if authors:
        joins.append("authors")
        filters.append(Author.id.in_(authors))
    if issn:
        joins.append("source")
        filters.append(Source.issn.ilike(f"%{issn.strip()}%"))
    if source_type:
        filters.append(Source.type == source_type)
    if citations_min is not None:
        filters.append(Publication.citations_count >= citations_min)
    if citations_max is not None:
        filters.append(Publication.citations_count <= citations_max)
    if percentile_min is not None:
        filters.append(Publication.percentile_2024 >= percentile_min)
    if percentile_max is not None:
        filters.append(Publication.percentile_2024 <= percentile_max)
    if q:
        q_like = f"%{q.strip()}%"
        joins.append("authors")
        filters.append(or_(Publication.title.ilike(q_like), Publication.doi.ilike(q_like), Author.display_name.ilike(q_like)))

    need_join_authors = "authors" in joins
    need_join_source = True
    stmt = select(Source.id, Source.name, func.count(func.distinct(Publication.id))).join(Publication.source)
    if need_join_authors:
        stmt = stmt.join(Publication.authors)
    if filters:
        stmt = stmt.where(and_(*filters))
    if query:
        stmt = stmt.where(Source.name.ilike(f"%{query.strip()}%"))
    stmt = stmt.group_by(Source.id, Source.name).order_by(desc(func.count(func.distinct(Publication.id)))).limit(limit)
    rows = db.execute(stmt).all()
    return [{"id": r[0], "name": r[1], "count": r[2]} for r in rows]


@router.get("/export")
def export_csv(
    q: Optional[str] = Query(default=None),
    year_min: Optional[int] = None,
    year_max: Optional[int] = None,
    quartiles: Optional[List[str]] = None,
    authors: Optional[List[int]] = None,
    sources: Optional[List[int]] = None,
    issn: Optional[str] = None,
    source_type: Optional[str] = None,
    citations_min: Optional[int] = None,
    citations_max: Optional[int] = None,
    percentile_min: Optional[int] = None,
    percentile_max: Optional[int] = None,
    sort: str = "year_desc",
    fmt: str = Query(default="csv", description="csv|xlsx"),
    db: Session = Depends(get_db),
):
    filters = [Publication.status == "approved"]
    joins: List[str] = []
    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)
    if quartiles:
        filters.append(Publication.quartile.in_(quartiles))
    if sources:
        filters.append(Publication.source_id.in_(sources))
    if source_type:
        joins.append("source")
        filters.append(Source.type == source_type)
    if issn:
        joins.append("source")
        filters.append(Source.issn.ilike(f"%{issn.strip()}%"))
    if citations_min is not None:
        filters.append(Publication.citations_count >= citations_min)
    if citations_max is not None:
        filters.append(Publication.citations_count <= citations_max)
    if percentile_min is not None:
        filters.append(Publication.percentile_2024 >= percentile_min)
    if percentile_max is not None:
        filters.append(Publication.percentile_2024 <= percentile_max)
    if q:
        q_like = f"%{q.strip()}%"
        joins += ["authors", "source"]
        filters.append(or_(Publication.title.ilike(q_like), Publication.doi.ilike(q_like), Source.name.ilike(q_like), Author.display_name.ilike(q_like)))
    if authors:
        joins.append("authors")
        filters.append(Author.id.in_(authors))

    stmt = select(Publication).options(joinedload(Publication.source), joinedload(Publication.authors))
    need_join_authors = "authors" in joins
    need_join_source = "source" in joins
    if need_join_authors:
        stmt = stmt.join(Publication.authors)
    if need_join_source:
        stmt = stmt.join(Publication.source)
    if filters:
        stmt = stmt.where(and_(*filters))

    if sort == "year_asc":
        stmt = stmt.order_by(asc(Publication.year), Publication.id)
    elif sort == "citations_desc":
        stmt = stmt.order_by(desc(Publication.citations_count), desc(Publication.year), Publication.id)
    elif sort == "citations_asc":
        stmt = stmt.order_by(asc(Publication.citations_count), desc(Publication.year), Publication.id)
    elif sort == "title_asc":
        stmt = stmt.order_by(asc(Publication.title), Publication.id)
    elif sort == "title_desc":
        stmt = stmt.order_by(desc(Publication.title), Publication.id)
    else:
        stmt = stmt.order_by(desc(Publication.year), Publication.id)

    rows = db.execute(stmt).scalars().unique().all()

    # CSV export
    if fmt.lower() == "csv":
        def iter_csv():
            header = [
                "id","year","title","authors","source","issn","doi","scopus_url","citations","quartile","percentile_2024","pdf_url"
            ]
            yield ",".join(header) + "\n"
            for p in rows:
                authors_str = "; ".join([a.display_name for a in p.authors])
                source_name = p.source.name if p.source else ""
                issn_val = p.source.issn if p.source else ""
                vals = [
                    str(p.id), str(p.year or ''), p.title.replace('"','""'), authors_str.replace('"','""'), source_name.replace('"','""'),
                    str(issn_val or ''), str(p.doi or ''), str(p.scopus_url or ''), str(p.citations_count or 0), str(p.quartile or ''), str(p.percentile_2024 or ''), str(p.pdf_url or ''),
                ]
                # Quote fields that may contain commas
                def q(v: str) -> str:
                    return f'"{v}"' if ("," in v or "\n" in v or '"' in v) else v
                yield ",".join([q(v) for v in vals]) + "\n"
        return StreamingResponse(iter_csv(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=export.csv"})

    # XLSX export
    if fmt.lower() == "xlsx":
        if Workbook is None:
            # Fallback to CSV if openpyxl not installed
            return StreamingResponse((line for line in []), media_type="text/plain", headers={"Content-Disposition": "attachment; filename=error.txt"})
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        header = ["id","year","title","authors","source","issn","doi","scopus_url","citations","quartile","percentile_2024","pdf_url"]
        ws.append(header)
        for p in rows:
            authors_str = "; ".join([a.display_name for a in p.authors])
            source_name = p.source.name if p.source else ""
            issn_val = p.source.issn if p.source else ""
            ws.append([
                p.id, p.year or '', p.title, authors_str, source_name,
                issn_val or '', p.doi or '', p.scopus_url or '', p.citations_count or 0, p.quartile or '', p.percentile_2024 or '', p.pdf_url or ''
            ])
        # Autosize columns (simple)
        for col in ws.columns:
            max_len = 10
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=export.xlsx"})

    # default CSV fallback
    def iter_fallback():
        yield "Unsupported format"
    return StreamingResponse(iter_fallback(), media_type="text/plain")


@router.get("/stats")
def stats(
    q: Optional[str] = Query(default=None),
    year_min: Optional[int] = None,
    year_max: Optional[int] = None,
    quartiles: Optional[List[str]] = None,
    authors: Optional[List[int]] = None,
    sources: Optional[List[int]] = None,
    issn: Optional[str] = None,
    source_type: Optional[str] = None,
    citations_min: Optional[int] = None,
    citations_max: Optional[int] = None,
    percentile_min: Optional[int] = None,
    percentile_max: Optional[int] = None,
    db: Session = Depends(get_db),
):
    filters = [Publication.status == "approved"]
    joins: List[str] = []
    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)
    if quartiles:
        filters.append(Publication.quartile.in_(quartiles))
    if sources:
        filters.append(Publication.source_id.in_(sources))
    if source_type:
        joins.append("source")
        filters.append(Source.type == source_type)
    if issn:
        joins.append("source")
        filters.append(Source.issn.ilike(f"%{issn.strip()}%"))
    if citations_min is not None:
        filters.append(Publication.citations_count >= citations_min)
    if citations_max is not None:
        filters.append(Publication.citations_count <= citations_max)
    if percentile_min is not None:
        filters.append(Publication.percentile_2024 >= percentile_min)
    if percentile_max is not None:
        filters.append(Publication.percentile_2024 <= percentile_max)
    if q:
        q_like = f"%{q.strip()}%"
        joins += ["authors", "source"]
        filters.append(or_(Publication.title.ilike(q_like), Publication.doi.ilike(q_like), Source.name.ilike(q_like), Author.display_name.ilike(q_like)))
    if authors:
        joins.append("authors")
        filters.append(Author.id.in_(authors))

    need_join_authors = "authors" in joins
    need_join_source = "source" in joins

    # KPI
    kpi_stmt = select(func.count(func.distinct(Publication.id)))
    if need_join_authors:
        kpi_stmt = kpi_stmt.join(Publication.authors)
    if need_join_source:
        kpi_stmt = kpi_stmt.join(Publication.source)
    if filters:
        kpi_stmt = kpi_stmt.where(and_(*filters))
    total_pubs = db.execute(kpi_stmt).scalar_one()

    authors_count_stmt = select(func.count(func.distinct(Author.id))).join(Publication.authors)
    if need_join_source:
        authors_count_stmt = authors_count_stmt.join(Publication.source)
    if filters:
        authors_count_stmt = authors_count_stmt.where(and_(*filters))
    total_authors = db.execute(authors_count_stmt).scalar_one()

    sources_count_stmt = select(func.count(func.distinct(Source.id))).join(Publication.source)
    if need_join_authors:
        sources_count_stmt = sources_count_stmt.join(Publication.authors)
    if filters:
        sources_count_stmt = sources_count_stmt.where(and_(*filters))
    total_sources = db.execute(sources_count_stmt).scalar_one()

    avg_per_author = (total_pubs / total_authors) if total_authors else 0.0

    # Yearly series
    year_stmt = select(Publication.year, func.count(Publication.id), func.sum(Publication.citations_count)).group_by(Publication.year).order_by(Publication.year)
    if need_join_authors:
        year_stmt = year_stmt.join(Publication.authors)
    if need_join_source:
        year_stmt = year_stmt.join(Publication.source)
    if filters:
        year_stmt = year_stmt.where(and_(*filters))
    year_rows = db.execute(year_stmt).all()
    yearly = [{"year": r[0], "publications": r[1], "citations": int(r[2] or 0)} for r in year_rows]

    # Top authors
    top_auth_stmt = select(Author.display_name, func.count(func.distinct(Publication.id))).join(Publication.authors)
    if need_join_source:
        top_auth_stmt = top_auth_stmt.join(Publication.source)
    if filters:
        top_auth_stmt = top_auth_stmt.where(and_(*filters))
    top_auth_stmt = top_auth_stmt.group_by(Author.display_name).order_by(desc(func.count(func.distinct(Publication.id)))).limit(10)
    top_authors = [{"author": r[0], "count": r[1]} for r in db.execute(top_auth_stmt).all()]

    # Top sources
    top_src_stmt = select(Source.name, func.count(func.distinct(Publication.id))).join(Publication.source)
    if need_join_authors:
        top_src_stmt = top_src_stmt.join(Publication.authors)
    if filters:
        top_src_stmt = top_src_stmt.where(and_(*filters))
    top_src_stmt = top_src_stmt.group_by(Source.name).order_by(desc(func.count(func.distinct(Publication.id)))).limit(10)
    top_sources = [{"source": r[0], "count": r[1]} for r in db.execute(top_src_stmt).all()]

    # Quartile distribution
    quart_stmt = select(Publication.quartile, func.count(Publication.id)).group_by(Publication.quartile)
    if need_join_authors:
        quart_stmt = quart_stmt.join(Publication.authors)
    if need_join_source:
        quart_stmt = quart_stmt.join(Publication.source)
    if filters:
        quart_stmt = quart_stmt.where(and_(*filters))
    quart = [{"quartile": (r[0] or "-"), "count": r[1]} for r in db.execute(quart_stmt).all()]

    return {
        "kpi": {"publications": total_pubs, "authors": total_authors, "sources": total_sources, "avg_per_author": avg_per_author},
        "yearly": yearly,
        "top_authors": top_authors,
        "top_sources": top_sources,
        "quartiles": quart,
    }
