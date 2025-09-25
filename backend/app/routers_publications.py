from __future__ import annotations

from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, Body, UploadFile, File, Form, Header
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, func, case
import os
import shutil

from .db import get_db
from .models import Publication, Author, Source, User, publication_authors
from .schemas import PublicationCreate, ValidateSourceResponse, PublicationOut
from .config import get_settings

# Initialize router early (must be defined before any @router.* usage)
router = APIRouter(prefix="/publications", tags=["publications"])

# -----------------------------
# Department → Faculty mapping (manual mapping provided by university)
# -----------------------------
# Note: keys must match User.department values exactly (case-sensitive strings from the database).
DEPT_TO_FAC: dict[str, str] = {
    # Шетел тілдері факультеті
    "Ағылшын және неміс тілдері кафедрасы": "Шетел тілдері факультеті",
    "Аударма ісі кафедрасы": "Шетел тілдері факультеті",
    "Шетел филологиясы кафедрасы": "Шетел тілдері факультеті",
    "Әлем тілдері кафедрасы": "Шетел тілдері факультеті",
    "Шетел филологиясы және аударма ісі кафедрасы": "Шетел тілдері факультеті",

    # Физика-математика факультеті
    "Информатика және есептеу техникасы кафедрасы": "Физика-математика факультеті",
    "Ақпараттық жүйелер кафедрасы": "Физика-математика факультеті",
    "Конденсацияланған күй физикасы кафедрасы": "Физика-математика факультеті",
    "Эксперименттік және теориялық физика кафедрасы": "Физика-математика факультеті",
    "Математика кафедрасы": "Физика-математика факультеті",
    "Негізгі және қолданбалы математика кафедрасы": "Физика-математика факультеті",
    "Информатика теориясы және оқыту технологиялары кафедрасы": "Физика-математика факультеті",
    "Информатика және ақпараттық технологиялар кафедрасы": "Физика-математика факультеті",
    "Физика кафедрасы": "Физика-математика факультеті",

    # Техникалық факультет
    "Мұнай-газ ісі кафедрасы": "Техникалық факультет",
    "Дизайн кафедрасы": "Техникалық факультет",
    "Металлургия және тау-кен ісі кафедрасы": "Техникалық факультет",
    "Автомобиль көлігі және жол қозғалысын ұйымдастыру кафедрасы": "Техникалық факультет",
    "Жалпы техникалық пәндер кафедрасы": "Техникалық факультет",
    "Химиялық технология кафедрасы": "Техникалық факультет",
    "Көлік техникасы, тасымалдауды ұйымдастыру және құрылыс кафедрасы": "Техникалық факультет",

    # Экономика және құқық факультеті
    "Қаржы және есеп кафедрасы": "Экономика және құқық факультеті",
    "Экономика және менеджмент кафедрасы": "Экономика және құқық факультеті",
    "Мемлекеттік басқару, қаржы және маркетинг кафедрасы": "Экономика және құқық факультеті",
    "Юриспруденция кафедрасы": "Экономика және құқық факультеті",
    "Мемлекеттік-құқықтық пәндер кафедрасы": "Экономика және құқық факультеті",

    # Жаратылыстану факультеті
    "Экология кафедрасы": "Жаратылыстану факультеті",
    "Химия және химиялық технология кафедрасы": "Жаратылыстану факультеті",
    "Биология кафедрасы": "Жаратылыстану факультеті",

    # Тарих факультеті
    "Тарих және аймақтану кафедрасы": "Тарих факультеті",
    "Философия кафедрасы": "Тарих факультеті",
    "Қазақстан тарихы және тарихи пәндер кафедрасы": "Тарих факультеті",
    "География және туризм кафедрасы": "Тарих факультеті",
    "Қазақстан халқы ассамблеясы және әлеуметтік-саяси пәндер кафедрасы": "Тарих факультеті",

    # Педагогикалық факультет
    "Теориялық және қолданбалы психология кафедрасы": "Педагогикалық факультет",
    "Әлеуметтік педагогика және бастауыш оқыту кафедрасы": "Педагогикалық факультет",
    "Мектепке дейінгі және арнайы білім беру кафедрасы": "Педагогикалық факультет",
    "Педагогика және білім психологиясы кафедрасы": "Педагогикалық факультет",
    "Психологиялық-педагогикалық және арнайы білім беру кафедрасы": "Педагогикалық факультет",
    "Педагогика, психология және бастауыш оқыту кафедрасы": "Педагогикалық факультет",

    # Филология факультеті
    "Орыс тілі мен әдебиеті кафедрасы": "Филология факультеті",
    "Қазақ әдебиеті кафедрасы": "Филология факультеті",
    "Қазақ тілінің теориялық және қолданбалы тіл білімі кафедрасы": "Филология факультеті",
    "Қазақ филологиясы кафедрасы": "Филология факультеті",
    "Орыс филологиясы және мәдениетаралық коммуникация кафедрасы": "Филология факультеті",

    # Кәсіби-шығармашылық факультет
    "Дене мәдениетінің теориялық негіздері кафедрасы": "Кәсіби-шығармашылық факультет",
    "Музыкалық білім кафедрасы": "Кәсіби-шығармашылық факультет",
    "Бейнелеу өнері және кәсіби оқыту кафедрасы": "Кәсіби-шығармашылық факультет",
    "Көркем еңбек және дизайн кафедрасы": "Кәсіби-шығармашылық факультет",
    "Музыка және хореография кафедрасы": "Кәсіби-шығармашылық факультет",
    "Дене тәрбиесі теориясы мен әдістемесі кафедрасы": "Кәсіби-шығармашылық факультет",
    "Дене тәрбиесі кафедрасы": "Кәсіби-шығармашылық факультет",
}

UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'uploads'))
_DEPT_MAP_CACHE: dict[str, str] | None = None
_DEPT_MAP_MTIME: float | None = None
_FACULTY_SET_CACHE: set[str] | None = None
_DEPT_NORM_MAP: dict[str, str] | None = None

def _norm_dept(s: str) -> str:
    x = (s or '').replace('\u00A0', ' ')
    # remove quotes and angle-quotes
    for ch in ['«','»','“','”','"','\'']:
        x = x.replace(ch, '')
    x = ' '.join(x.strip().split())
    x_low = x.lower()
    # remove trailing 'кафедра'/'кафедрасы'
    for suf in [' кафедрасы', ' кафедра']:
        if x_low.endswith(suf):
            x_low = x_low[: -len(suf)]
            break
    return x_low.strip()

def _load_dept_map() -> dict[str, str]:
    global _DEPT_MAP_CACHE, _DEPT_MAP_MTIME, _DEPT_NORM_MAP
    try:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        path = os.path.join(UPLOAD_DIR, "_dept_map.json")
        if os.path.isfile(path):
            mtime = os.path.getmtime(path)
            if _DEPT_MAP_CACHE is None or _DEPT_MAP_MTIME != mtime:
                import json
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                # expect { department: faculty }
                if isinstance(data, dict):
                    _DEPT_MAP_CACHE = {str(k): str(v) for k, v in data.items()}
                    _DEPT_MAP_MTIME = mtime
                    # invalidate faculty set cache
                    globals()["_FACULTY_SET_CACHE"] = None
        if _DEPT_MAP_CACHE is None:
            _DEPT_MAP_CACHE = dict(DEPT_TO_FAC)
        # Build normalized lookup map alongside
        _DEPT_NORM_MAP = { _norm_dept(k): v for k, v in _DEPT_MAP_CACHE.items() }
        return _DEPT_MAP_CACHE
    except Exception:
        # fallback to built-in
        return dict(DEPT_TO_FAC)

def _faculty_set() -> set[str]:
    global _FACULTY_SET_CACHE
    if _FACULTY_SET_CACHE is None:
        m = _load_dept_map()
        _FACULTY_SET_CACHE = set(m.values())
    return _FACULTY_SET_CACHE

def map_faculty(department: str | None, user_faculty: str | None) -> str:
    # Prefer explicit mapping by department
    if department:
        m = _load_dept_map()
        # exact match first
        fac = m.get(department)
        if not fac and _DEPT_NORM_MAP is not None:
            fac = _DEPT_NORM_MAP.get(_norm_dept(department))
        if fac:
            return fac
    # Accept user_faculty only if it is a known faculty name
    if user_faculty and user_faculty in _faculty_set():
        return user_faculty
    return 'Без привязки'


# -----------------------------
# Doc type normalization helpers
# -----------------------------
def _norm_doc_type_expr(col):
    # lower(trim(col)) without NBSPs
    from sqlalchemy import func
    return func.lower(func.trim(func.replace(col, '\u00A0', ' ')))

def _doc_type_matches(col, value: str):
    # Match by normalized equality OR substring (to be tolerant to variants)
    from sqlalchemy import or_, literal
    v = (value or '').strip().lower()
    if not v:
        return literal(True)
    norm_col = _norm_doc_type_expr(col)
    return or_(norm_col == v, norm_col.ilike(f"%{v}%"))


@router.get("/doc_types")
def list_doc_types(
    upload_source: Optional[str] = Query(default=None, description="kokson|scopus|manual"),
    db: Session = Depends(get_db),
):
    """Return distinct non-empty doc_type values for populating UI filters."""
    from sqlalchemy import select
    stmt = select(Publication.doc_type).where(Publication.doc_type.is_not(None))
    if upload_source:
        stmt = stmt.where(Publication.upload_source == upload_source)
    rows = db.execute(stmt).all()
    vals = sorted({(r[0] or '').strip() for r in rows if (r[0] or '').strip()})
    return {"items": vals}


@router.get("/dept_map")
def get_dept_map():
    """Return department -> faculty mapping currently in use (merged built-in + overrides).
    Frontend can use this to filter departments by faculty.
    """
    try:
        m = _load_dept_map()
        return {"map": m}
    except Exception:
        return {"map": dict(DEPT_TO_FAC)}


def _role_from_token(token: str) -> str | None:
    s = get_settings()
    if token == s.ADMIN_TOKEN:
        return "admin"
    if token == s.USER_TOKEN:
        return "user"
    return None


def require_uploader(
    authorization: str | None = Header(default=None),
    x_client_id: str | None = Header(default=None),
    x_user_id: str | None = Header(default=None),
) -> tuple[str, str, int | None]:
    """Return (role, client_id, user_id). Only allow user/admin tokens.
    x_user_id is optional and may be a stringifiable int from frontend login.
    """
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    token = authorization.split()[-1]
    role = _role_from_token(token)
    if not role:
        raise HTTPException(status_code=401, detail="Invalid token")
    if not x_client_id:
        # Auto-generate a client id for Swagger/manual calls to avoid ASCII header issues
        try:
            import uuid
            x_client_id = f"auto-{uuid.uuid4()}"
        except Exception:
            x_client_id = "auto-client"
    try:
        uid = int(x_user_id) if (x_user_id is not None and str(x_user_id).strip() != '') else None
    except Exception:
        uid = None
    return role, x_client_id, uid

# Resolve uploads directory from settings (supports Render Disk via UPLOAD_DIR env)
_settings = get_settings()
UPLOAD_DIR = _settings.UPLOAD_DIR
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("/validate/source", response_model=ValidateSourceResponse)
def validate_source(
    issn: Optional[str] = Query(default=None),
    name: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _dep: tuple[str, str, int | None] = Depends(require_uploader),
):
    if not issn and not name:
        raise HTTPException(status_code=400, detail="Provide issn or name")

    stmt = select(Source)
    if issn:
        stmt = stmt.where(Source.issn == issn)
    elif name:
        stmt = stmt.where(Source.name.ilike(name))

    src = db.execute(stmt).scalars().first()
    if src:
        return ValidateSourceResponse(found=True, source=src)  # type: ignore[arg-type]
    return ValidateSourceResponse(found=False, message="Источник не найден в базе")


@router.get("/mine", response_model=list[PublicationOut])
def list_my_publications(
    db: Session = Depends(get_db),
    actor: tuple[str, str, int | None] = Depends(require_uploader),
):
    """Return publications visible to the current user.
    Rules:
    - Admin: all items.
    - Non-admin: union of
        a) items uploaded by this client (uploader_id == X-Client-Id), and
        b) items linked to this user_id OR matched by author-name to user's full name
           (strict normalized equality + fallback by last name with initials + name variants).
    """
    from sqlalchemy.orm import selectinload
    from sqlalchemy import case as _case, or_ as _or, select as _select

    role, client_id, user_id = actor

    # Base query builder
    stmt = select(Publication).options(
        selectinload(Publication.source),
        selectinload(Publication.authors),
    )

    if role != "admin":
        conds = [Publication.uploader_id == client_id]
        # Keep only explicit user linkage to avoid heavy scans
        if user_id is not None:
            conds.append(Publication.user_id == user_id)

        from sqlalchemy import or_ as _or
        stmt = stmt.where(_or(*conds))
        # Show all statuses for personal view, including 'pending'
        # (no status filter for non-admin)

    # Order: rejected first (need fix), then pending (on moderation), then approved. Then by year desc, id
    # Order: rejected first, then approved; admins may still have 'pending', which will go last
    status_order = _case(
        (Publication.status == 'rejected', 0),
        (Publication.status == 'approved', 1),
        else_=2,
    )
    stmt = stmt.order_by(status_order, Publication.year.desc(), Publication.id)
    rows = db.execute(stmt).scalars().all()
    return [PublicationOut.model_validate(r) for r in rows]


@router.post("", response_model=PublicationOut, status_code=201)
def create_publication(payload: PublicationCreate, db: Session = Depends(get_db)):
    # Source detect or create (lightweight)
    src: Optional[Source] = None
    if payload.issn:
        src = db.execute(select(Source).where(Source.issn == payload.issn)).scalars().first()
    if not src and payload.source_name:
        src = db.execute(select(Source).where(Source.name.ilike(payload.source_name))).scalars().first()
    if not src and (payload.issn or payload.source_name):
        src = Source(
            name=payload.source_name or (payload.issn or "Unknown"),
            issn=payload.issn,
            type=(payload.source_type or "journal"),
        )
        db.add(src)
        db.flush()

    # Create or reuse authors
    author_objs: List[Author] = []
    for idx, name in enumerate(payload.authors):
        nm = (name or "").strip()
        if not nm:
            continue
        existing = db.execute(select(Author).where(Author.display_name == nm)).scalars().first()
        if existing:
            author_objs.append(existing)
        else:
            a = Author(display_name=nm, normalized_name=nm.lower())
            db.add(a)
            db.flush()
            author_objs.append(a)

    pub = Publication(
        title=payload.title.strip(),
        year=payload.year,
        doi=(payload.doi or None),
        pdf_url=(payload.pdf_url or None),
        citations_count=payload.citations_count or 0,
        quartile=None,  # may be inferred from src later
        source=src,
        status="pending",  # submissions go to moderation
    )
    # Любое редактирование через личный кабинет переводит публикацию в pending на повторную модерацию
    # (админ правит через админ-роуты; здесь всегда pending)
    pub.status = "pending"

    db.add(pub)
    db.flush()

    # link authors preserving order
    pub.authors = author_objs

    # infer quartile from source if present
    if src and src.sjr_quartile and not pub.quartile:
        pub.quartile = src.sjr_quartile

    db.commit()
    db.refresh(pub)
    return pub  # type: ignore[return-value]


@router.post("/upload", status_code=201)
async def upload_publication(
    title: str = Form(...),
    year: int = Form(...),
    authors: str = Form(...),  # semicolon-separated list (main authors)
    coauthors: Optional[str] = Form(None),  # semicolon-separated list (coauthors)
    source_name: Optional[str] = Form(None),
    issn: Optional[str] = Form(None),
    doi: Optional[str] = Form(None),
    citations_count: Optional[int] = Form(0),
    quartile: Optional[str] = Form(None),
    percentile_2024: Optional[int] = Form(None),
    language: Optional[str] = Form(None),
    url: Optional[str] = Form(None),
    scopus_url: Optional[str] = Form(None),
    upload_source: Optional[str] = Form("scopus"),  # scopus|article (legacy 'kokson')|manual
    doc_type: Optional[str] = Form(None),
    published_date: Optional[str] = Form(None),  # dd.mm.yyyy or yyyy-mm-dd
    user_id: Optional[int] = Form(None),  # only admin may set
    user_login: Optional[str] = Form(None),  # admin convenience: link by login
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    actor: tuple[str, str, int | None] = Depends(require_uploader),
):
    # Normalize legacy alias
    us = (upload_source or "scopus").lower()
    if us == "kokson":
        us = "article"

    # 1) Save file to uploads/ (optional for Scopus; required for Article is validated below)
    filename: Optional[str] = None
    if file is not None:
        safe_title = "_".join(title.strip().split())[:60]
        filename = f"{safe_title}_{year}_{file.filename}"
        file_path = os.path.join(UPLOAD_DIR, filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

    # 2) Source detect or create (by strict name first)
    src: Optional[Source] = None
    if source_name:
        src = db.execute(select(Source).where(Source.name == source_name.strip())).scalars().first()
    if not src and issn:
        src = db.execute(select(Source).where(Source.issn == issn.strip())).scalars().first()
    if not src and (source_name or issn):
        src = Source(
            name=source_name or (issn or "Unknown"),
            issn=issn,
            type="journal",
        )
        db.add(src)
        db.flush()

    # 3) Authors and Coauthors (split by ';')
    main_list = [a.strip() for a in authors.split(';') if a.strip()]
    co_list = [a.strip() for a in (coauthors or '').split(';') if a.strip()]
    combined = main_list + co_list
    author_objs: List[Author] = []
    for nm in combined:
        existing = db.execute(select(Author).where(Author.display_name == nm)).scalars().first()
        if existing:
            author_objs.append(existing)
        else:
            a = Author(display_name=nm, normalized_name=" ".join(nm.lower().split()))
            db.add(a)
            db.flush()
            author_objs.append(a)

    # 4) Create publication (pending)
    role, client_id, _user_id = actor

    # For Article uploads enforce required URL and PDF
    if us == "article":
        if not url or not url.strip():
            raise HTTPException(status_code=400, detail="URL is required for Article uploads")
        if not file:
            raise HTTPException(status_code=400, detail="PDF file is required for Article uploads")
    # Derive year from published_date if provided
    parsed_date = None
    if published_date:
        try:
            # dd.mm.yyyy
            d, m, y = published_date.strip().split('.')
            parsed_date = (int(y), int(m), int(d))
        except Exception:
            try:
                # yyyy-mm-dd
                y, m, d = published_date.strip().split('-')
                parsed_date = (int(y), int(m), int(d))
            except Exception:
                parsed_date = None

    # If parsed_date ok, override year
    if parsed_date:
        try:
            from datetime import date
            _y, _m, _d = parsed_date
            pub_date = date(_y, _m, _d)
            year = _y
        except Exception:
            pub_date = None
    else:
        pub_date = None

    pub = Publication(
        title=title.strip(),
        year=year if not pub_date else year,  # year already set above from parsed_date
        doi=(doi or None),
        pdf_url=(f"/uploads/{filename}" if filename else None),  # relative URL served by static
        url=(url or None),
        scopus_url=(scopus_url or None),
        citations_count=citations_count or 0,
        quartile=quartile,
        percentile_2024=percentile_2024,
        source=src,
        status="pending",
        uploader_id=client_id,
        uploaded_by_role=role,
        language=(language or None),
        upload_source=us,
        doc_type=(doc_type or None),
        published_date=pub_date,
        main_authors_count=(len(main_list) or None),
    )
    if role == "admin":
        if user_id:
            pub.user_id = user_id
        elif (user_login or "").strip():
            from .models import User
            from sqlalchemy import func
            u = db.execute(select(User).where(User.login == user_login.strip())).scalar_one_or_none()
            if u:
                pub.user_id = u.id
                fac = func.coalesce(User.faculty, 'Без привязки')
    db.add(pub)
    db.flush()
    pub.authors = author_objs
    db.commit()
    db.refresh(pub)
    return {"message": "Publication uploaded", "publication_id": pub.id, "pdf_url": pub.pdf_url}


# Public Kokson listing for UI (table/cards on /kokso)
@router.get("/kokson", response_model=List[PublicationOut])
def list_kokson(
    q: Optional[str] = Query(default=None),
    year_min: Optional[int] = Query(default=None),
    year_max: Optional[int] = Query(default=None),
    issn: Optional[str] = Query(default=None),
    lang: Optional[str] = Query(default=None, description="ru|kz|en"),
    doc_type: Optional[str] = Query(default=None),
    authors: Optional[List[int]] = Query(default=None),
    status: Optional[str] = Query(default=None, description="pending|approved|rejected|all"),
    faculty: Optional[str] = Query(default=None, description="Filter by mapped faculty; use 'Без привязки' for unlinked"),
    sort: str = Query(default="year_desc", description="year_desc|year_asc|title_asc|title_desc|type_asc|type_desc|author_asc|author_desc"),
    db: Session = Depends(get_db),
):
    from sqlalchemy.orm import selectinload
    from sqlalchemy import and_, or_, asc, desc
    stmt = (
        select(Publication)
        .options(selectinload(Publication.source), selectinload(Publication.authors))
    )
    # Accept both new 'article' and legacy 'kokson'
    filters = [Publication.upload_source.in_(["article", "kokson"])]
    if status in ("pending","approved","rejected"):
        filters.append(Publication.status == status)
    joins = []
    if year_min is not None:
        filters.append(Publication.year >= year_min)
    if year_max is not None:
        filters.append(Publication.year <= year_max)
    if lang:
        filters.append(Publication.language == lang)
    if doc_type:
        from sqlalchemy import literal
        filters.append(func.lower(func.trim(Publication.doc_type)) == func.lower(func.trim(literal(doc_type))))
    if issn:
        joins.append("source")
        from .models import Source as Src
        # Join via relationship to avoid cartesian products
        stmt = stmt.join(Publication.source)
        filters.append(Src.issn.ilike(f"%{issn.strip()}%"))
    if authors:
        joins.append("authors")
        from .models import Author as Au
        stmt = stmt.join(Publication.authors)
        filters.append(Au.id.in_(authors))
    if q:
        joins += ["authors", "source"]
        from .models import Source as Src, Author as Au
        if "authors" not in joins:
            stmt = stmt.join(Publication.authors)
        if "source" not in joins:
            # Join via relationship to avoid cartesian products
            stmt = stmt.join(Publication.source)
        q_like = f"%{q.strip()}%"
        filters.append(or_(Publication.title.ilike(q_like), Src.name.ilike(q_like), Au.display_name.ilike(q_like)))

    if filters:
        stmt = stmt.where(and_(*filters))

    # Faculty filtering
    if faculty and faculty.strip():
        fac_val = faculty.strip()
        if fac_val == 'Без привязки':
            # Materialize candidate rows and filter in Python based on mapped faculty of MAIN authors only
            from sqlalchemy import outerjoin, literal
            from .models import User as U
            # Join main authors (author_order < COALESCE(main_authors_count, 999))
            j = (
                outerjoin(Publication, publication_authors, publication_authors.c.publication_id == Publication.id)
                .join(Author, Author.id == publication_authors.c.author_id)
                .outerjoin(U, U.id == Author.user_id)
            )
            base = (
                select(
                    Publication.id,
                    publication_authors.c.author_order,
                    func.coalesce(Publication.main_authors_count, literal(999)).label("main_cnt"),
                    U.department,
                    U.faculty,
                )
                .select_from(j)
                .where(publication_authors.c.author_order < func.coalesce(Publication.main_authors_count, literal(999)))
            )
            rows = db.execute(base).all()
            by_pub: dict[int, list[tuple[str|None,str|None]]] = {}
            for pid, ao, main_cnt, dept, ufac in rows:
                by_pub.setdefault(int(pid), []).append((dept, ufac))
            ids: list[int] = []
            # Decide: include pub if all MAIN authors map to 'Без привязки' OR there are no linked main authors
            for pid, pairs in by_pub.items():
                if not pairs:
                    ids.append(pid); continue
                all_unlinked = True
                for dept, ufac in pairs:
                    if map_faculty(dept, ufac) != 'Без привязки':
                        all_unlinked = False
                        break
                if all_unlinked:
                    ids.append(pid)
            if not ids:
                return []
            # IMPORTANT: keep previously applied filters (upload_source='kokson', year/doc_type/etc)
            # Add ID filter instead of recreating a fresh select
            stmt = stmt.where(Publication.id.in_(ids))
        else:
            from .models import User as U
            stmt = stmt.join(U, U.id == Publication.user_id).where(U.faculty == fac_val)

    if sort == "year_asc":
        stmt = stmt.order_by(asc(Publication.year), Publication.id)
    elif sort == "title_asc":
        stmt = stmt.order_by(asc(Publication.title), Publication.id)
    elif sort == "title_desc":
        stmt = stmt.order_by(desc(Publication.title), Publication.id)
    elif sort == "type_asc":
        stmt = stmt.order_by(asc(Publication.doc_type.nullslast()), Publication.id)
    elif sort == "type_desc":
        stmt = stmt.order_by(desc(Publication.doc_type.nullslast()), Publication.id)
    elif sort == "author_asc":
        # order by primary author's display_name (lexicographically)
        from .models import Author as Au
        stmt = stmt.join(Publication.authors).order_by(asc(Au.display_name), Publication.id)
    elif sort == "author_desc":
        from .models import Author as Au
        stmt = stmt.join(Publication.authors).order_by(desc(Au.display_name), Publication.id)
    else:
        stmt = stmt.order_by(desc(Publication.year), Publication.id)

    rows = db.execute(stmt).scalars().unique().all()
    return [PublicationOut.model_validate(p) for p in rows]

@router.api_route("/mine/{pub_id}", methods=["PATCH", "POST"], response_model=PublicationOut)
async def update_my_publication(
    pub_id: int,
    title: Optional[str] = Form(None),
    year: Optional[int] = Form(None),
    doi: Optional[str] = Form(None),
    citations_count: Optional[int] = Form(None),
    quartile: Optional[str] = Form(None),
    percentile_2024: Optional[int] = Form(None),
    url: Optional[str] = Form(None),
    language: Optional[str] = Form(None),
    doc_type: Optional[str] = Form(None),
    note: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    actor: tuple[str, str, int | None] = Depends(require_uploader),
    x_user_id: Optional[int] = Header(default=None, alias="X-User-Id"),
):
    role, client_id, hdr_user_id = actor
    pub = db.get(Publication, pub_id)
    if not pub:
        raise HTTPException(status_code=404, detail="Publication not found")
    # Ownership: admin OR same uploader client OR same user id linked via Publication.user_id or Author.user_id
    user_id = x_user_id or hdr_user_id
    is_owner = False
    if role == "admin":
        is_owner = True
    if pub.uploader_id and pub.uploader_id == client_id:
        is_owner = True
    if user_id and pub.user_id and pub.user_id == user_id:
        is_owner = True
    if not is_owner and user_id:
        authors = db.execute(
            select(Author)
            .join(publication_authors, publication_authors.c.author_id == Author.id)
            .where(publication_authors.c.publication_id == pub.id)
        ).scalars().all()
        is_owner = any(a.user_id == user_id for a in authors)
    if not is_owner:
        raise HTTPException(status_code=403, detail="Forbidden")

    # Replace file if provided (optional)
    if file is not None:
        # remove old file
        try:
            if pub.pdf_url:
                old = os.path.join(UPLOAD_DIR, os.path.basename(pub.pdf_url))
                if os.path.isfile(old):
                    os.remove(old)
        except Exception:
            pass
        safe_title = "_".join((title or pub.title).strip().split())[:60]
        original_name = os.path.basename(file.filename)
        filename = f"{safe_title}_{(year or pub.year)}_{original_name}"
        file_path = os.path.join(UPLOAD_DIR, filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        pub.pdf_url = f"/uploads/{filename}"

    if title is not None:
        pub.title = title
    if year is not None:
        pub.year = year
    if doi is not None:
        pub.doi = doi
    if citations_count is not None:
        pub.citations_count = citations_count
    if quartile is not None:
        pub.quartile = quartile
    if percentile_2024 is not None:
        pub.percentile_2024 = percentile_2024
    if url is not None:
        pub.url = url or None
    if language is not None:
        pub.language = language or None
    if doc_type is not None:
        pub.doc_type = doc_type or None
    if note is not None:
        pub.note = note or None

    # After any personal edit, send back to moderation
    pub.status = "pending"
    pub.note = None

    db.add(pub)
    db.commit()
    db.refresh(pub)
    return PublicationOut.model_validate(pub)


# -----------------------------
# Articles (Kokson) Statistics
# -----------------------------

@router.get("/stats/articles/language_share")
def stats_articles_language_share(
    year_min: Optional[int] = Query(default=None),
    year_max: Optional[int] = Query(default=None),
    faculty: Optional[str] = Query(default=None),
    doc_type: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Return per-year language shares for Kokson-approved publications.
    If faculty is provided, only publications linked to users of that faculty are counted.
    Otherwise, include all publications (even without user linkage).
    Output: [{ year, total, ru, kz, en, ru_pct, kz_pct, en_pct }]
    """
    from .models import User

    # Special case: faculty == 'Без привязки' — compute in Python by mapping departments
    if faculty and faculty.strip() == 'Без привязки':
        # LEFT JOIN to users to extract dept/fac; then map per-row and aggregate only "Без привязки"
        from sqlalchemy import outerjoin
        j = outerjoin(Publication, User, User.id == Publication.user_id)
        q = select(
            Publication.year,
            Publication.language,
            User.department,
            User.faculty,
        ).select_from(j).where(Publication.upload_source == "kokson", Publication.status == "approved")
        if year_min is not None:
            q = q.where(Publication.year >= year_min)
        if year_max is not None:
            q = q.where(Publication.year <= year_max)
        if doc_type:
            from sqlalchemy import literal
            q = q.where(func.lower(func.trim(Publication.doc_type)) == func.lower(func.trim(literal(doc_type))))
        q = q.order_by(Publication.year)
        rows = db.execute(q).all()
        agg: dict[int, dict[str, int]] = {}
        for year, lang, dept, ufac in rows:
            mapped = map_faculty(dept, ufac)
            if mapped != 'Без привязки':
                continue
            bucket = agg.setdefault(int(year or 0), {"total": 0, "ru": 0, "kz": 0, "en": 0})
            bucket["total"] += 1
            if (lang or '').lower() == 'ru': bucket["ru"] += 1
            elif (lang or '').lower() == 'kz': bucket["kz"] += 1
            elif (lang or '').lower() == 'en': bucket["en"] += 1
        out = []
        for year in sorted(agg.keys()):
            total = int(agg[year]["total"])
            ru = int(agg[year]["ru"]); kz = int(agg[year]["kz"]); en = int(agg[year]["en"])
            if total <= 0:
                ru_pct = kz_pct = en_pct = 0.0
            else:
                ru_pct = round(ru * 100.0 / total, 2)
                kz_pct = round(kz * 100.0 / total, 2)
                en_pct = round(en * 100.0 / total, 2)
            out.append({
                "year": year,
                "total": total,
                "ru": ru,
                "kz": kz,
                "en": en,
                "ru_pct": ru_pct,
                "kz_pct": kz_pct,
                "en_pct": en_pct,
            })
        return out


@router.post("/mine/{pub_id}")
def update_my_publication(
    pub_id: int,
    title: Optional[str] = Form(default=None),
    year: Optional[int] = Form(default=None),
    doi: Optional[str] = Form(default=None),
    url: Optional[str] = Form(default=None),
    citations_count: Optional[int] = Form(default=None),
    quartile: Optional[str] = Form(default=None),
    percentile_2024: Optional[int] = Form(default=None),
    file: Optional[UploadFile] = File(default=None, description="Optional PDF/Word file"),
    db: Session = Depends(get_db),
    x_user_id: Optional[int] = Header(default=None, alias="X-User-Id"),
):
    """Allow a logged-in user to update their own publication. All fields are optional.
    A user is considered owner if (a) Publication.user_id == X-User-Id, or
    (b) any linked Author has author.user_id == X-User-Id.
    """
    user_id = x_user_id
    if not user_id:
        raise HTTPException(status_code=401, detail="Missing user id")
    pub = db.get(Publication, pub_id)
    if not pub:
        raise HTTPException(status_code=404, detail="Publication not found")
    # ownership check
    is_owner = (pub.user_id == user_id)
    if not is_owner:
        # join authors lazily
        try:
            authors = db.execute(select(Author).join(publication_authors, publication_authors.c.author_id == Author.id).where(publication_authors.c.publication_id == pub.id)).scalars().all()
            is_owner = any(a.user_id == user_id for a in authors)
        except Exception:
            is_owner = False
    if not is_owner:
        raise HTTPException(status_code=403, detail="Not your publication")

    # Update provided fields only
    if title is not None:
        pub.title = title
    if year is not None:
        pub.year = year
    if doi is not None:
        pub.doi = doi
    if url is not None:
        pub.url = url
    if citations_count is not None:
        pub.citations_count = citations_count
    if quartile is not None:
        pub.quartile = quartile
    if percentile_2024 is not None:
        pub.percentile_2024 = percentile_2024

    # Optional file upload
    if file is not None:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        safe_name = f"pub_{pub.id}_{file.filename}"
        safe_name = safe_name.replace("/", "_").replace("\\", "_")
        path = os.path.join(UPLOAD_DIR, safe_name)
        with open(path, "wb") as out:
            shutil.copyfileobj(file.file, out)
        pub.pdf_url = f"/uploads/{safe_name}"

    db.add(pub)
    db.commit()
    db.refresh(pub)
    return PublicationOut.model_validate(pub)

    # Default fast SQL path (no faculty or a specific known faculty)
    q = select(
        Publication.year,
        func.count(Publication.id).label("total"),
        func.sum(case((Publication.language == 'ru', 1), else_=0)).label("ru"),
        func.sum(case((Publication.language == 'kz', 1), else_=0)).label("kz"),
        func.sum(case((Publication.language == 'en', 1), else_=0)).label("en"),
    ).where(Publication.upload_source == "kokson")
    if year_min is not None:
        q = q.where(Publication.year >= year_min)
    if year_max is not None:
        q = q.where(Publication.year <= year_max)
    if doc_type:
        q = q.where(_doc_type_matches(Publication.doc_type, doc_type))
    if faculty:
        q = q.join(User, User.id == Publication.user_id).where(User.faculty == faculty)
    q = q.group_by(Publication.year).order_by(Publication.year)
    rows = db.execute(q).all()
    out = []
    for year, total, ru, kz, en in rows:
        total = int(total or 0)
        ru = int(ru or 0); kz = int(kz or 0); en = int(en or 0)
        if total <= 0:
            ru_pct = kz_pct = en_pct = 0.0
        else:
            ru_pct = round(ru * 100.0 / total, 2)
            kz_pct = round(kz * 100.0 / total, 2)
            en_pct = round(en * 100.0 / total, 2)
        out.append({
            "year": year,
            "total": total,
            "ru": ru,
            "kz": kz,
            "en": en,
            "ru_pct": ru_pct,
            "kz_pct": kz_pct,
            "en_pct": en_pct,
        })
    return out


_FACDEP_CACHE: dict | None = None
_FACDEP_MTIME: float | None = None

def _load_facdep() -> dict:
    """Load faculties/departments from aku_faculties_departments.xlsx if present.
    Expected columns (case-insensitive): Faculty, Department.
    Returns { faculties: [...], departments: [...], map: { department: faculty } }.
    Cached with mtime to avoid reread."""
    import json
    global _FACDEP_CACHE, _FACDEP_MTIME
    # Look in project root and backend dir
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    candidates = [
        os.path.join(root_dir, "aku_faculties_departments.xlsx"),
        os.path.join(backend_dir, "aku_faculties_departments.xlsx"),
    ]
    path = next((p for p in candidates if os.path.isfile(p)), None)
    if not path:
        # Fallback to dept_map (merged) if excel absent
        m = _load_dept_map()
        facs = sorted({v for v in m.values() if v})
        deps = sorted({k for k in m.keys() if k})
        return {"faculties": facs, "departments": deps, "map": m}
    try:
        mtime = os.path.getmtime(path)
        if _FACDEP_CACHE is not None and _FACDEP_MTIME == mtime:
            return _FACDEP_CACHE
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            # If openpyxl not available, return fallback
            m = _load_dept_map()
            facs = sorted({v for v in m.values() if v})
            deps = sorted({k for k in m.keys() if k})
            return {"faculties": facs, "departments": deps, "map": m}
        wb = load_workbook(path)
        ws = wb.active
        # Find columns
        header = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
        def col_index(names: list[str]) -> int | None:
            low = [h.lower() for h in header]
            for i, h in enumerate(low):
                if h in names:
                    return i
            return None
        ci_fac = col_index(["faculty", "факультет" ])
        ci_dep = col_index(["department", "кафедра" ])
        facs_set: set[str] = set()
        deps_set: set[str] = set()
        mp: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2):
            fac = str(row[ci_fac].value).strip() if (ci_fac is not None and row[ci_fac].value is not None) else ''
            dep = str(row[ci_dep].value).strip() if (ci_dep is not None and row[ci_dep].value is not None) else ''
            if not fac and not dep:
                continue
            if fac:
                facs_set.add(fac)
            if dep:
                deps_set.add(dep)
            if dep and fac:
                mp[dep] = fac
        facs = sorted(facs_set)
        deps = sorted(deps_set)
        data = {"faculties": facs, "departments": deps, "map": mp}
        _FACDEP_CACHE = data; _FACDEP_MTIME = mtime
        return data
    except Exception:
        m = _load_dept_map()
        facs = sorted({v for v in m.values() if v})
        deps = sorted({k for k in m.keys() if k})
        return {"faculties": facs, "departments": deps, "map": m}


@router.get("/facdep")
def get_facdep():
    """Return faculties/departments from aku_faculties_departments.xlsx (if present),
    otherwise use internal mapping and Users facets fallback."""
    return _load_facdep()

@router.get("/stats/language_share")
def stats_language_share_all(
    year_min: Optional[int] = Query(default=None),
    year_max: Optional[int] = Query(default=None),
    doc_type: Optional[str] = Query(default=None, description="Filter by doc_type across ALL approved publications"),
    db: Session = Depends(get_db),
):
    """Return per-year language shares for ALL approved publications (any upload_source).
    Optional filter by doc_type. Output is the same shape as /stats/articles/language_share.
    """
    q = select(
        Publication.year,
        func.count(Publication.id).label("total"),
        func.sum(case((Publication.language == 'ru', 1), else_=0)).label("ru"),
        func.sum(case((Publication.language == 'kz', 1), else_=0)).label("kz"),
        func.sum(case((Publication.language == 'en', 1), else_=0)).label("en"),
    )
    if year_min is not None:
        q = q.where(Publication.year >= year_min)
    if year_max is not None:
        q = q.where(Publication.year <= year_max)
    if doc_type:
        from sqlalchemy import literal
        q = q.where(func.lower(func.trim(Publication.doc_type)) == func.lower(func.trim(literal(doc_type))))
    q = q.group_by(Publication.year).order_by(Publication.year)
    rows = db.execute(q).all()
    out = []
    for year, total, ru, kz, en in rows:
        total = int(total or 0)
        ru = int(ru or 0); kz = int(kz or 0); en = int(en or 0)
        if total <= 0:
            ru_pct = kz_pct = en_pct = 0.0
        else:
            ru_pct = round(ru * 100.0 / total, 2)
            kz_pct = round(kz * 100.0 / total, 2)
            en_pct = round(en * 100.0 / total, 2)
        out.append({
            "year": year,
            "total": total,
            "ru": ru,
            "kz": kz,
            "en": en,
            "ru_pct": ru_pct,
            "kz_pct": kz_pct,
            "en_pct": en_pct,
        })
    return out

@router.get("/stats/articles/language_share/export")
def export_stats_articles_language_share(
    year_min: Optional[int] = Query(default=None),
    year_max: Optional[int] = Query(default=None),
    doc_type: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Export Articles language share to XLSX with filename 'stati_zh.xlsx'."""
    from fastapi.responses import Response
    data = stats_articles_language_share(year_min=year_min, year_max=year_max, doc_type=doc_type, db=db)
    headers_row = ["year","total","ru","kz","en","ru_pct","kz_pct","en_pct"]
    try:
        from openpyxl import Workbook
        import io
        wb = Workbook(); ws = wb.active
        ws.title = "language_share"
        ws.append(headers_row)
        for r in data:
            ws.append([
                r.get("year"), r.get("total"), r.get("ru"), r.get("kz"), r.get("en"), r.get("ru_pct"), r.get("kz_pct"), r.get("en_pct")
            ])
        bio = io.BytesIO(); wb.save(bio); bio.seek(0)
        return Response(
            content=bio.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=stati_zh.xlsx"},
        )
    except Exception:
        # Fallback to CSV if openpyxl unavailable
        import io, csv
        buf = io.StringIO(); w = csv.writer(buf)
        w.writerow(headers_row)
        for r in data:
            w.writerow([r.get("year"), r.get("total"), r.get("ru"), r.get("kz"), r.get("en"), r.get("ru_pct"), r.get("kz_pct"), r.get("en_pct")])
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=stati_zh.csv"},
        )

@router.get("/stats/articles/faculty_summary")
def stats_articles_faculty_summary(
    year: Optional[int] = Query(default=None),
    doc_type: Optional[str] = Query(default=None),
    authors: Optional[List[int]] = Query(default=None, description="Filter by Author IDs (Kokson authors)"),
    db: Session = Depends(get_db),
):
    """Return counts per faculty for a given year and optional doc_type.
    Faculty is derived from department via DEPT_TO_FAC when possible.
    Publications without user linkage appear under 'Без привязки'.
    """
    from .models import User
    stmt = (
        select(User.faculty, User.department, func.count().label("count"))
        .select_from(Publication)
        .join(User, User.id == Publication.user_id, isouter=True)
        .where(Publication.upload_source == "kokson")
    )
    if authors:
        # Join only MAIN authors: author_order < COALESCE(main_authors_count, 999)
        from sqlalchemy import literal, and_, text
        stmt = (
            stmt.join(publication_authors, publication_authors.c.publication_id == Publication.id)
                .join(Author, Author.id == publication_authors.c.author_id)
                .where(Author.id.in_(authors))
                .where(publication_authors.c.author_order < func.coalesce(Publication.main_authors_count, literal(999)))
        )
    stmt = stmt.group_by(User.faculty, User.department)
    if year is not None:
        stmt = stmt.where(Publication.year == year)
    if doc_type:
        stmt = stmt.where(_doc_type_matches(Publication.doc_type, doc_type))
    rows = db.execute(stmt).all()
    agg: dict[str, int] = {}
    for fac, dept, cnt in rows:
        key = map_faculty(dept, fac)
        agg[key] = agg.get(key, 0) + int(cnt or 0)
    return [{"faculty": k, "count": v} for k, v in sorted(agg.items(), key=lambda x: x[0])]


@router.get("/stats/articles/faculty_breakdown")
def stats_articles_faculty_breakdown(
    year: Optional[int] = Query(default=None),
    faculty: str = Query(...),
    doc_type: Optional[str] = Query(default=None),
    authors: Optional[List[int]] = Query(default=None, description="Filter by Author IDs (Kokson authors)"),
    db: Session = Depends(get_db),
):
    """Return counts per department for a given faculty/year/doc_type.
    Faculty is determined by DEPT_TO_FAC mapping.
    """
    from .models import User
    stmt = (
        select(User.department.label("department"), User.faculty, func.count().label("count"))
        .select_from(Publication)
        .join(User, User.id == Publication.user_id)
        .where(Publication.upload_source == "kokson")
    )
    if authors:
        stmt = stmt.join(Publication.authors).where(Author.id.in_(authors))
    stmt = stmt.group_by(User.department, User.faculty)
    if year is not None:
        stmt = stmt.where(Publication.year == year)
    if doc_type:
        stmt = stmt.where(_doc_type_matches(Publication.doc_type, doc_type))
    rows = db.execute(stmt).all()
    out = []
    for dept, fac, cnt in rows:
        mapped = map_faculty(dept, fac)
        if mapped == faculty:
            out.append({"department": dept or "—", "count": int(cnt or 0)})
    out.sort(key=lambda r: r["department"])
    return out


@router.get("/stats/articles/faculty_summary/export")
def export_stats_articles_faculty_summary(
    year: int = Query(...),
    doc_type: Optional[str] = Query(default=None),
    fmt: str = Query(default="xlsx", description="xlsx|csv"),
    db: Session = Depends(get_db),
):
    from fastapi.responses import Response
    data = stats_articles_faculty_summary(year=year, doc_type=doc_type, db=db)
    headers_row = ["Факультет", "Количество"]
    rows = [[r["faculty"], r["count"]] for r in data]
    if fmt == "csv":
        import io, csv
        buf = io.StringIO(); w = csv.writer(buf)
        w.writerow(headers_row); w.writerows(rows)
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=articles_faculty_{year}.csv"},
        )
    # xlsx
    try:
        from openpyxl import Workbook
        import io
        wb = Workbook(); ws = wb.active
        ws.append(headers_row)
        for r in rows: ws.append(r)
        bio = io.BytesIO(); wb.save(bio); bio.seek(0)
        return Response(
            content=bio.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=articles_faculty_{year}.xlsx"},
        )
    except Exception:
        import io, csv
        buf = io.StringIO(); w = csv.writer(buf)
        w.writerow(headers_row); w.writerows(rows)
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=articles_faculty_{year}.csv"},
        )
